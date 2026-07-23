"""
HOJAS 4-6: Top rotación, Análisis Clientes, Temporal & Estacionalidad
"""
import pandas as pd
import numpy as np
import warnings
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

df = pd.read_pickle('df_clean.pkl')
prod_agg = pd.read_pickle('prod_agg.pkl')
top20_fact = pd.read_pickle('top20_fact.pkl')

MAX_DATE = df['Fecha'].max()
MIN_DATE = df['Fecha'].min()

FONT_NAME = 'Arial'
COLOR_HEADER = '1F4E78'
COLOR_SUBHEADER = '5B9BD5'
COLOR_TOTAL = 'D9E1F2'
COLOR_ALERT_BG = 'F8CBAD'
COLOR_GROW = 'C6EFCE'
COLOR_DROP = 'FFC7CE'

thin = Side(border_style='thin', color='BFBFBF')
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

def fmt_header(cell):
    cell.font = Font(name=FONT_NAME, size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', start_color=COLOR_HEADER)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_all

def fmt_title(cell):
    cell.font = Font(name=FONT_NAME, size=14, bold=True, color=COLOR_HEADER)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def fmt_total_row(cell):
    cell.font = Font(name=FONT_NAME, size=10, bold=True)
    cell.fill = PatternFill('solid', start_color=COLOR_TOTAL)
    cell.border = border_all

def fmt_cell(cell, num_fmt=None, bold=False):
    cell.font = Font(name=FONT_NAME, size=10, bold=bold)
    cell.border = border_all
    if num_fmt:
        cell.number_format = num_fmt

def autosize(ws, widths):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

CURRENCY = '$#,##0;($#,##0);-'
INT_FMT = '#,##0;(#,##0);-'
PCT_FMT = '0.0%;(0.0%);-'

wb = load_workbook('Analisis_Global_Family_Mayo2026.xlsx')
total_global = df['Total'].sum()
unidades_tot = df['Cantidad'].sum()

# =========================================================
# HOJA 4: TOP 20 POR ROTACIÓN
# =========================================================
print('Construyendo HOJA 4: Top 20 por rotación...')
ws = wb.create_sheet('4. Top20 por Rotacion')
ws['A1'] = 'TOP 20 PRODUCTOS POR ROTACIÓN (unidades vendidas)'
fmt_title(ws['A1'])
ws.merge_cells('A1:I1')

top20_rot = prod_agg.nlargest(20, 'Unidades').reset_index(drop=True)

headers = ['#', 'Producto', 'Categoría', 'Proveedor', 'Unidades', '% Unid. Total', 'Ingresos', 'Precio Prom.', 'Nº Facturas', 'Nº Clientes']
for i, h in enumerate(headers):
    c = ws.cell(row=3, column=i+1, value=h); fmt_header(c)

row = 4
start_row = row
for i, r in top20_rot.iterrows():
    ws.cell(row=row, column=1, value=i+1)
    ws.cell(row=row, column=2, value=r['Descripcion Producto'])
    ws.cell(row=row, column=3, value=r['CategoriaProducto'])
    ws.cell(row=row, column=4, value=r['Proveedor'])
    ws.cell(row=row, column=5, value=int(r['Unidades']))
    ws.cell(row=row, column=6, value=f'=E{row}/{int(unidades_tot)}')
    ws.cell(row=row, column=7, value=int(r['Ingresos']))
    ws.cell(row=row, column=8, value=f'=IFERROR(G{row}/E{row},0)')
    ws.cell(row=row, column=9, value=int(r['Facturas']))
    ws.cell(row=row, column=10, value=int(r['Clientes']))
    for col in range(1, 11):
        fmt_cell(ws.cell(row=row, column=col))
    ws.cell(row=row, column=5).number_format = INT_FMT
    ws.cell(row=row, column=6).number_format = PCT_FMT
    ws.cell(row=row, column=7).number_format = CURRENCY
    ws.cell(row=row, column=8).number_format = CURRENCY
    ws.cell(row=row, column=9).number_format = INT_FMT
    ws.cell(row=row, column=10).number_format = INT_FMT
    row += 1

ws.cell(row=row, column=2, value='TOTAL TOP 20')
ws.cell(row=row, column=5, value=f'=SUM(E{start_row}:E{row-1})')
ws.cell(row=row, column=6, value=f'=SUM(F{start_row}:F{row-1})')
ws.cell(row=row, column=7, value=f'=SUM(G{start_row}:G{row-1})')
ws.cell(row=row, column=9, value=f'=SUM(I{start_row}:I{row-1})')
ws.cell(row=row, column=10, value=f'=SUM(J{start_row}:J{row-1})')
for col in range(1, 11):
    fmt_total_row(ws.cell(row=row, column=col))
ws.cell(row=row, column=5).number_format = INT_FMT
ws.cell(row=row, column=6).number_format = PCT_FMT
ws.cell(row=row, column=7).number_format = CURRENCY
ws.cell(row=row, column=9).number_format = INT_FMT
ws.cell(row=row, column=10).number_format = INT_FMT

autosize(ws, {'A': 4, 'B': 55, 'C': 24, 'D': 16, 'E': 12, 'F': 12, 'G': 18, 'H': 14, 'I': 12, 'J': 12})

top20_rot.to_pickle('top20_rot.pkl')

# =========================================================
# HOJA 5: ANÁLISIS DE CLIENTES
# =========================================================
print('Construyendo HOJA 5: Análisis de Clientes...')
ws = wb.create_sheet('5. Analisis Clientes')
ws['A1'] = 'ANÁLISIS DE CLIENTES'
fmt_title(ws['A1'])
ws.merge_cells('A1:I1')

# Agregación por cliente (todos, incluyendo CONSUMIDOR FINAL)
cli_agg = df.groupby('Cliente').agg(
    Ingresos=('Total', 'sum'),
    Facturas=('Factura', 'nunique'),
    Unidades=('Cantidad', 'sum'),
    FechaMin=('Fecha', 'min'),
    FechaMax=('Fecha', 'max'),
    Generico=('ClienteGenerico', 'first')
).reset_index()
cli_agg['Ticket'] = cli_agg['Ingresos'] / cli_agg['Facturas']
cli_agg['DiasDesdeUltima'] = (MAX_DATE - cli_agg['FechaMax']).dt.days
cli_agg['DiasActivo'] = (cli_agg['FechaMax'] - cli_agg['FechaMin']).dt.days + 1
cli_agg['FrecPromedio'] = cli_agg.apply(
    lambda r: r['DiasActivo'] / r['Facturas'] if r['Facturas'] > 0 else np.nan, axis=1)

# Cliente -> categoría top
cat_por_cliente = df.groupby(['Cliente', 'CategoriaProducto'])['Total'].sum().reset_index()
cat_top = cat_por_cliente.sort_values(['Cliente', 'Total'], ascending=[True, False]).groupby('Cliente').head(3)
cat_top_str = cat_top.groupby('Cliente')['CategoriaProducto'].apply(lambda x: ' | '.join(x)).to_dict()
cli_agg['CategTop3'] = cli_agg['Cliente'].map(cat_top_str)

# Top 20 clientes (excluyendo genéricos)
cli_real = cli_agg[~cli_agg['Generico']].copy()
top20_cli = cli_real.nlargest(20, 'Ingresos').reset_index(drop=True)

# ---- Subhoja: Top 20 ----
ws['A3'] = 'TOP 20 CLIENTES POR FACTURACIÓN (excluye CONSUMIDOR FINAL)'
ws['A3'].font = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_HEADER)

headers = ['#', 'Cliente', 'Facturación', '% Total', 'Nº Facturas', 'Unidades', 'Ticket Prom.', 'Primera Compra', 'Última Compra', 'Días desde últ.', 'Frec. compra (días)', 'Top 3 categorías']
for i, h in enumerate(headers):
    c = ws.cell(row=5, column=i+1, value=h); fmt_header(c)

row = 6
start_row = row
for i, r in top20_cli.iterrows():
    ws.cell(row=row, column=1, value=i+1)
    ws.cell(row=row, column=2, value=r['Cliente'])
    ws.cell(row=row, column=3, value=int(r['Ingresos']))
    ws.cell(row=row, column=4, value=f'=C{row}/{int(total_global)}')
    ws.cell(row=row, column=5, value=int(r['Facturas']))
    ws.cell(row=row, column=6, value=int(r['Unidades']))
    ws.cell(row=row, column=7, value=f'=IFERROR(C{row}/E{row},0)')
    ws.cell(row=row, column=8, value=r['FechaMin'].date())
    ws.cell(row=row, column=9, value=r['FechaMax'].date())
    ws.cell(row=row, column=10, value=int(r['DiasDesdeUltima']))
    ws.cell(row=row, column=11, value=round(r['FrecPromedio'], 1))
    ws.cell(row=row, column=12, value=r['CategTop3'])
    for col in range(1, 13):
        fmt_cell(ws.cell(row=row, column=col))
    ws.cell(row=row, column=3).number_format = CURRENCY
    ws.cell(row=row, column=4).number_format = PCT_FMT
    ws.cell(row=row, column=5).number_format = INT_FMT
    ws.cell(row=row, column=6).number_format = INT_FMT
    ws.cell(row=row, column=7).number_format = CURRENCY
    ws.cell(row=row, column=8).number_format = 'yyyy-mm-dd'
    ws.cell(row=row, column=9).number_format = 'yyyy-mm-dd'
    ws.cell(row=row, column=10).number_format = INT_FMT
    ws.cell(row=row, column=11).number_format = '0.0'
    row += 1

# Total top 20
ws.cell(row=row, column=2, value='TOTAL TOP 20')
ws.cell(row=row, column=3, value=f'=SUM(C{start_row}:C{row-1})')
ws.cell(row=row, column=4, value=f'=SUM(D{start_row}:D{row-1})')
ws.cell(row=row, column=5, value=f'=SUM(E{start_row}:E{row-1})')
ws.cell(row=row, column=6, value=f'=SUM(F{start_row}:F{row-1})')
for col in range(1, 13):
    fmt_total_row(ws.cell(row=row, column=col))
ws.cell(row=row, column=3).number_format = CURRENCY
ws.cell(row=row, column=4).number_format = PCT_FMT
ws.cell(row=row, column=5).number_format = INT_FMT
ws.cell(row=row, column=6).number_format = INT_FMT

# ---- Segmentación ABC ----
row += 3
ws.cell(row=row, column=1, value='SEGMENTACIÓN ABC POR FACTURACIÓN (todos los clientes reales)').font = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_HEADER)
row += 1
cli_real_sorted = cli_real.sort_values('Ingresos', ascending=False).reset_index(drop=True)
n_total = len(cli_real_sorted)
n_a = max(1, int(round(n_total * 0.2)))
n_b = max(1, int(round(n_total * 0.3)))
seg_a = cli_real_sorted.iloc[:n_a]
seg_b = cli_real_sorted.iloc[n_a:n_a+n_b]
seg_c = cli_real_sorted.iloc[n_a+n_b:]

headers = ['Segmento', 'Definición', '# Clientes', '% Clientes', 'Facturación', '% Facturación', 'Ticket Prom. cliente']
for i, h in enumerate(headers):
    c = ws.cell(row=row, column=i+1, value=h); fmt_header(c)
row += 1
seg_start = row

for label, definition, seg in [
    ('A', 'Top 20% por facturación', seg_a),
    ('B', 'Siguiente 30%', seg_b),
    ('C', 'Restante 50%', seg_c)
]:
    n = len(seg)
    ing = seg['Ingresos'].sum()
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=definition)
    ws.cell(row=row, column=3, value=n)
    ws.cell(row=row, column=4, value=f'=C{row}/{n_total}')
    ws.cell(row=row, column=5, value=int(ing))
    ws.cell(row=row, column=6, value=f'=E{row}/{int(cli_real_sorted["Ingresos"].sum())}')
    ws.cell(row=row, column=7, value=f'=IFERROR(E{row}/C{row},0)')
    for col in range(1, 8):
        fmt_cell(ws.cell(row=row, column=col))
    ws.cell(row=row, column=3).number_format = INT_FMT
    ws.cell(row=row, column=4).number_format = PCT_FMT
    ws.cell(row=row, column=5).number_format = CURRENCY
    ws.cell(row=row, column=6).number_format = PCT_FMT
    ws.cell(row=row, column=7).number_format = CURRENCY
    row += 1

ws.cell(row=row, column=1, value='TOTAL')
ws.cell(row=row, column=3, value=f'=SUM(C{seg_start}:C{row-1})')
ws.cell(row=row, column=4, value=f'=SUM(D{seg_start}:D{row-1})')
ws.cell(row=row, column=5, value=f'=SUM(E{seg_start}:E{row-1})')
ws.cell(row=row, column=6, value=f'=SUM(F{seg_start}:F{row-1})')
for col in range(1, 8):
    fmt_total_row(ws.cell(row=row, column=col))
ws.cell(row=row, column=3).number_format = INT_FMT
ws.cell(row=row, column=4).number_format = PCT_FMT
ws.cell(row=row, column=5).number_format = CURRENCY
ws.cell(row=row, column=6).number_format = PCT_FMT

# ---- Alerta de clientes inactivos ----
row += 3
ws.cell(row=row, column=1, value='ALERTA RECUPERACIÓN — clientes activos antes que NO compran hace 60+ días').font = Font(name=FONT_NAME, size=12, bold=True, color='C00000')
row += 1
# Activos antes: que tengan al menos 3 facturas en su historia y que su última compra sea hace 60+ días
inactivos = cli_real[
    (cli_real['Facturas'] >= 3) &
    (cli_real['DiasDesdeUltima'] >= 60)
].sort_values('Ingresos', ascending=False).reset_index(drop=True)

headers = ['#', 'Cliente', 'Facturación histórica', 'Nº Facturas', 'Última compra', 'Días sin comprar', 'Frec. promedio (días)']
for i, h in enumerate(headers):
    c = ws.cell(row=row, column=i+1, value=h); fmt_header(c)
row += 1
inact_start = row
for i, r in inactivos.iterrows():
    ws.cell(row=row, column=1, value=i+1)
    ws.cell(row=row, column=2, value=r['Cliente'])
    ws.cell(row=row, column=3, value=int(r['Ingresos']))
    ws.cell(row=row, column=4, value=int(r['Facturas']))
    ws.cell(row=row, column=5, value=r['FechaMax'].date())
    ws.cell(row=row, column=6, value=int(r['DiasDesdeUltima']))
    ws.cell(row=row, column=7, value=round(r['FrecPromedio'], 1))
    for col in range(1, 8):
        fmt_cell(ws.cell(row=row, column=col))
    ws.cell(row=row, column=3).number_format = CURRENCY
    ws.cell(row=row, column=4).number_format = INT_FMT
    ws.cell(row=row, column=5).number_format = 'yyyy-mm-dd'
    ws.cell(row=row, column=6).number_format = INT_FMT
    ws.cell(row=row, column=7).number_format = '0.0'
    row += 1

# Formato condicional: pintar de rojo claro los días sin comprar > 90
if len(inactivos) > 0:
    ws.conditional_formatting.add(
        f'F{inact_start}:F{row-1}',
        CellIsRule(operator='greaterThan', formula=['90'], fill=PatternFill('solid', start_color=COLOR_DROP))
    )
    ws.conditional_formatting.add(
        f'F{inact_start}:F{row-1}',
        CellIsRule(operator='between', formula=['60', '90'], fill=PatternFill('solid', start_color=COLOR_ALERT_BG))
    )

autosize(ws, {'A': 5, 'B': 48, 'C': 18, 'D': 12, 'E': 14, 'F': 14, 'G': 14, 'H': 14, 'I': 14, 'J': 14, 'K': 14, 'L': 45})

cli_agg.to_pickle('cli_agg.pkl')
cli_real.to_pickle('cli_real.pkl')

# =========================================================
# HOJA 6: ANÁLISIS TEMPORAL Y ESTACIONALIDAD
# =========================================================
print('Construyendo HOJA 6: Análisis Temporal y Estacionalidad...')
ws = wb.create_sheet('6. Temporal Estacionalidad')

ws['A1'] = 'ANÁLISIS TEMPORAL Y ESTACIONALIDAD'
fmt_title(ws['A1'])
ws.merge_cells('A1:G1')

# Pivot mes x año (en este caso sobre todo Oct-2025 a May-2026, sin YoY pero estructura preparada)
df['MesNum'] = df['Fecha'].dt.month
df['MesNom'] = df['Fecha'].dt.strftime('%m-%b')
piv = df.pivot_table(index='MesNom', columns='Año', values='Total', aggfunc='sum').fillna(0)
piv = piv.sort_index()

ws['A3'] = 'Facturación mensual por año (filas = meses, columnas = años)'
ws['A3'].font = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_HEADER)

row = 5
ws.cell(row=row, column=1, value='Mes'); fmt_header(ws.cell(row=row, column=1))
for j, yr in enumerate(piv.columns):
    c = ws.cell(row=row, column=2+j, value=str(int(yr))); fmt_header(c)
ws.cell(row=row, column=2+len(piv.columns), value='Total'); fmt_header(ws.cell(row=row, column=2+len(piv.columns)))
row += 1
piv_start = row
for mes, r in piv.iterrows():
    ws.cell(row=row, column=1, value=mes)
    for j, yr in enumerate(piv.columns):
        v = r[yr]
        c = ws.cell(row=row, column=2+j, value=int(v) if v > 0 else 0)
        fmt_cell(c, num_fmt=CURRENCY)
    # Total
    last_col = get_column_letter(1+len(piv.columns))
    ws.cell(row=row, column=2+len(piv.columns), value=f'=SUM(B{row}:{last_col}{row})')
    fmt_cell(ws.cell(row=row, column=2+len(piv.columns)), num_fmt=CURRENCY)
    fmt_cell(ws.cell(row=row, column=1), bold=True)
    row += 1
# Total general
ws.cell(row=row, column=1, value='TOTAL')
for j in range(len(piv.columns)+1):
    col_l = get_column_letter(2+j)
    ws.cell(row=row, column=2+j, value=f'=SUM({col_l}{piv_start}:{col_l}{row-1})')
for col in range(1, 3+len(piv.columns)):
    fmt_total_row(ws.cell(row=row, column=col))
    if col >= 2:
        ws.cell(row=row, column=col).number_format = CURRENCY

# Mes pico y valle
row += 2
mes_total = df.groupby('AñoMes')['Total'].sum().sort_values(ascending=False)
ws.cell(row=row, column=1, value='Meses pico (top 3):').font = Font(name=FONT_NAME, size=11, bold=True)
row += 1
for am, v in mes_total.head(3).items():
    ws.cell(row=row, column=1, value=am)
    ws.cell(row=row, column=2, value=int(v))
    fmt_cell(ws.cell(row=row, column=1)); fmt_cell(ws.cell(row=row, column=2), num_fmt=CURRENCY)
    row += 1
row += 1
ws.cell(row=row, column=1, value='Meses valle (bottom 3):').font = Font(name=FONT_NAME, size=11, bold=True)
row += 1
for am, v in mes_total.tail(3).items():
    ws.cell(row=row, column=1, value=am)
    ws.cell(row=row, column=2, value=int(v))
    fmt_cell(ws.cell(row=row, column=1)); fmt_cell(ws.cell(row=row, column=2), num_fmt=CURRENCY)
    row += 1

# Día de semana
row += 2
ws.cell(row=row, column=1, value='FACTURACIÓN PROMEDIO POR DÍA DE LA SEMANA').font = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_HEADER)
row += 1
dia_map = {'Monday':'Lunes','Tuesday':'Martes','Wednesday':'Miércoles','Thursday':'Jueves','Friday':'Viernes','Saturday':'Sábado','Sunday':'Domingo'}
orden = ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado','Domingo']
df['DiaSemEs'] = df['DiaSemana'].map(dia_map)
dia_agg = df.groupby('DiaSemEs').agg(Total=('Total','sum'), DiasUnicos=('FechaSolo','nunique'))
dia_agg['Promedio'] = dia_agg['Total'] / dia_agg['DiasUnicos']
dia_agg = dia_agg.reindex(orden).dropna()

headers = ['Día', 'Facturación total', 'Días con ventas', 'Facturación promedio / día']
for i, h in enumerate(headers):
    c = ws.cell(row=row, column=i+1, value=h); fmt_header(c)
row += 1
for d, r in dia_agg.iterrows():
    ws.cell(row=row, column=1, value=d)
    ws.cell(row=row, column=2, value=int(r['Total']))
    ws.cell(row=row, column=3, value=int(r['DiasUnicos']))
    ws.cell(row=row, column=4, value=f'=IFERROR(B{row}/C{row},0)')
    fmt_cell(ws.cell(row=row, column=1), bold=True)
    fmt_cell(ws.cell(row=row, column=2), num_fmt=CURRENCY)
    fmt_cell(ws.cell(row=row, column=3), num_fmt=INT_FMT)
    fmt_cell(ws.cell(row=row, column=4), num_fmt=CURRENCY)
    row += 1
dia_max = dia_agg['Promedio'].idxmax()
dia_min = dia_agg['Promedio'].idxmin()
row += 1
ws.cell(row=row, column=1, value=f'Día con MAYOR facturación promedio: {dia_max} (${dia_agg.loc[dia_max,"Promedio"]:,.0f}/día)').font = Font(name=FONT_NAME, size=10, bold=True, color='006100')
row += 1
ws.cell(row=row, column=1, value=f'Día con MENOR facturación promedio: {dia_min} (${dia_agg.loc[dia_min,"Promedio"]:,.0f}/día)').font = Font(name=FONT_NAME, size=10, bold=True, color='9C0006')

# Tendencia últimos 6 meses
row += 3
ws.cell(row=row, column=1, value='TENDENCIA — últimos 6 meses (regresión lineal simple)').font = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_HEADER)
row += 1
mes_ord = df.groupby('AñoMes')['Total'].sum().sort_index().tail(6)
xs = np.arange(len(mes_ord))
ys = mes_ord.values
slope, intercept = np.polyfit(xs, ys, 1) if len(xs) >= 2 else (0, ys.mean() if len(ys) else 0)
# Calc tendencia %
mes_inicial = ys[0] if len(ys) else 0
mes_final = ys[-1] if len(ys) else 0
crec_pct = (mes_final / mes_inicial - 1) if mes_inicial > 0 else 0
if slope > 0 and crec_pct > 0.05:
    diagnostico = 'CRECIENTE'
    color_diag = '006100'
elif slope < 0 and crec_pct < -0.05:
    diagnostico = 'DECRECIENTE'
    color_diag = '9C0006'
else:
    diagnostico = 'ESTABLE'
    color_diag = 'BF8F00'

ws.cell(row=row, column=1, value='Mes')
ws.cell(row=row, column=2, value='Facturación')
fmt_header(ws.cell(row=row, column=1)); fmt_header(ws.cell(row=row, column=2))
row += 1
for am, v in mes_ord.items():
    ws.cell(row=row, column=1, value=am)
    ws.cell(row=row, column=2, value=int(v))
    fmt_cell(ws.cell(row=row, column=1)); fmt_cell(ws.cell(row=row, column=2), num_fmt=CURRENCY)
    row += 1
row += 1
ws.cell(row=row, column=1, value=f'Diagnóstico: {diagnostico}').font = Font(name=FONT_NAME, size=12, bold=True, color=color_diag)
row += 1
ws.cell(row=row, column=1, value=f'Pendiente (regresión lineal): {slope:,.0f} $/mes').font = Font(name=FONT_NAME, size=10, italic=True)
row += 1
ws.cell(row=row, column=1, value=f'Cambio mes inicial vs final: {crec_pct*100:+.1f}%').font = Font(name=FONT_NAME, size=10, italic=True)

autosize(ws, {'A': 30, 'B': 22, 'C': 18, 'D': 22, 'E': 18, 'F': 18, 'G': 18})

wb.save('Analisis_Global_Family_Mayo2026.xlsx')
print('Hojas 4-6 OK, archivo guardado.')
