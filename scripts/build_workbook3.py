"""
HOJAS 7-9: Productos Estratégicos, Concentración & Riesgo, Oportunidades Comerciales
"""
import pandas as pd
import numpy as np
import warnings
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from itertools import combinations
from collections import Counter

warnings.filterwarnings('ignore')

df = pd.read_pickle('df_clean.pkl')
prod_agg = pd.read_pickle('prod_agg.pkl')
top20_fact = pd.read_pickle('top20_fact.pkl')
top20_rot = pd.read_pickle('top20_rot.pkl')
cli_agg = pd.read_pickle('cli_agg.pkl')
cli_real = pd.read_pickle('cli_real.pkl')

MAX_DATE = df['Fecha'].max()
LAST_3M_START = MAX_DATE - pd.Timedelta(days=90)
PREV_3M_START = MAX_DATE - pd.Timedelta(days=180)

FONT_NAME = 'Arial'
COLOR_HEADER = '1F4E78'
COLOR_TOTAL = 'D9E1F2'
COLOR_ALERT_BG = 'F8CBAD'
COLOR_GROW = 'C6EFCE'
COLOR_DROP = 'FFC7CE'

thin = Side(border_style='thin', color='BFBFBF')
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

def fmt_header(cell):
    cell.font = Font(name=FONT_NAME, size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', start_color=COLOR_HEADER)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_all

def fmt_title(cell):
    cell.font = Font(name=FONT_NAME, size=14, bold=True, color=COLOR_HEADER)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def fmt_total_row(cell):
    cell.font = Font(name=FONT_NAME, size=10, bold=True)
    cell.fill = PatternFill('solid', start_color=COLOR_TOTAL)
    cell.border = border_all

def fmt_cell(cell, num_fmt=None, bold=False):
    cell.font = Font(name=FONT_NAME, size=10, bold=bold)
    cell.border = border_all
    if num_fmt:
        cell.number_format = num_fmt

def autosize(ws, widths):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

def fmt_subtitle(cell):
    cell.font = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_HEADER)

CURRENCY = '$#,##0;($#,##0);-'
INT_FMT = '#,##0;(#,##0);-'
PCT_FMT = '0.0%;(0.0%);-'

wb = load_workbook('Analisis_Global_Family_Mayo2026.xlsx')
total_global = df['Total'].sum()
unidades_tot = df['Cantidad'].sum()

# =========================================================
# HOJA 7: PRODUCTOS ESTRATÉGICOS
# =========================================================
print('Construyendo HOJA 7: Productos Estratégicos...')
ws = wb.create_sheet('7. Productos Estrategicos')
ws['A1'] = 'PRODUCTOS ESTRATÉGICOS — clasificación BCG simplificada'
fmt_title(ws['A1'])
ws.merge_cells('A1:G1')

ws['A2'] = ('Estrella = en Top 20 facturación Y Top 20 rotación. Vaca lechera = en Top 20 facturación pero NO en Top 20 rotación. '
            'Volumen = en Top 20 rotación pero NO en Top 20 facturación.')
ws['A2'].font = Font(name=FONT_NAME, size=9, italic=True, color='595959')
ws.merge_cells('A2:G2')

set_f = set(top20_fact['Cod Interno'])
set_r = set(top20_rot['Cod Interno'])
estrellas_ids = set_f & set_r
vacas_ids = set_f - set_r
volumen_ids = set_r - set_f

def build_block(ws, row, title, color, ids):
    ws.cell(row=row, column=1, value=title).font = Font(name=FONT_NAME, size=12, bold=True, color=color)
    row += 1
    headers = ['Producto', 'Categoría', 'Proveedor', 'Ingresos', 'Unidades', 'Precio Prom.', 'Nº Clientes']
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i+1, value=h); fmt_header(c)
    row += 1
    subset = prod_agg[prod_agg['Cod Interno'].isin(ids)].sort_values('Ingresos', ascending=False)
    for _, r in subset.iterrows():
        ws.cell(row=row, column=1, value=r['Descripcion Producto'])
        ws.cell(row=row, column=2, value=r['CategoriaProducto'])
        ws.cell(row=row, column=3, value=r['Proveedor'])
        ws.cell(row=row, column=4, value=int(r['Ingresos']))
        ws.cell(row=row, column=5, value=int(r['Unidades']))
        ws.cell(row=row, column=6, value=f'=IFERROR(D{row}/E{row},0)')
        ws.cell(row=row, column=7, value=int(r['Clientes']))
        for col in range(1, 8):
            fmt_cell(ws.cell(row=row, column=col))
        ws.cell(row=row, column=4).number_format = CURRENCY
        ws.cell(row=row, column=5).number_format = INT_FMT
        ws.cell(row=row, column=6).number_format = CURRENCY
        ws.cell(row=row, column=7).number_format = INT_FMT
        row += 1
    return row + 2

row = 4
row = build_block(ws, row, f'⭐ PRODUCTOS ESTRELLA ({len(estrellas_ids)})', '006100', estrellas_ids)
row = build_block(ws, row, f'🐄 PRODUCTOS VACA LECHERA — alto valor, baja rotación ({len(vacas_ids)})', 'BF8F00', vacas_ids)
row = build_block(ws, row, f'📦 PRODUCTOS VOLUMEN — alta rotación, bajo valor ({len(volumen_ids)})', '305496', volumen_ids)

# Productos en declive (últimos 3M vs 3M anteriores)
prod_3m = df[df['Fecha'] >= LAST_3M_START].groupby('Cod Interno').agg(
    Vent3M=('Total', 'sum'),
    Desc=('Descripcion Producto', 'first'),
    Cat=('CategoriaProducto', 'first')
)
prod_3m_prev = df[(df['Fecha'] >= PREV_3M_START) & (df['Fecha'] < LAST_3M_START)].groupby('Cod Interno').agg(
    Vent3MPrev=('Total', 'sum')
)
prod_trend = prod_3m.join(prod_3m_prev, how='outer').fillna(0)
# Solo consideramos productos con cierta materialidad: que hayan vendido al menos 500k en los últimos 6 meses
prod_trend['VentTotal6M'] = prod_trend['Vent3M'] + prod_trend['Vent3MPrev']
prod_trend_mat = prod_trend[prod_trend['VentTotal6M'] >= 500000].copy()
prod_trend_mat['Delta_Pct'] = (prod_trend_mat['Vent3M'] - prod_trend_mat['Vent3MPrev']) / prod_trend_mat['Vent3MPrev'].replace(0, np.nan)
prod_trend_mat['Delta_Abs'] = prod_trend_mat['Vent3M'] - prod_trend_mat['Vent3MPrev']

# Top 10 caídas (deben tener ventas previas > 0)
declive = prod_trend_mat[prod_trend_mat['Vent3MPrev'] > 0].sort_values('Delta_Pct').head(10)
# Top 10 ascensos
ascenso = prod_trend_mat[prod_trend_mat['Vent3MPrev'] > 0].sort_values('Delta_Pct', ascending=False).head(10)

def build_trend_block(ws, row, title, color, subset, drop=False):
    ws.cell(row=row, column=1, value=title).font = Font(name=FONT_NAME, size=12, bold=True, color=color)
    row += 1
    headers = ['Producto', 'Categoría', 'Ventas 3M previos', 'Ventas 3M recientes', 'Δ absoluta', 'Δ %']
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i+1, value=h); fmt_header(c)
    row += 1
    start = row
    for _, r in subset.iterrows():
        ws.cell(row=row, column=1, value=r['Desc'])
        ws.cell(row=row, column=2, value=r['Cat'])
        ws.cell(row=row, column=3, value=int(r['Vent3MPrev']))
        ws.cell(row=row, column=4, value=int(r['Vent3M']))
        ws.cell(row=row, column=5, value=f'=D{row}-C{row}')
        ws.cell(row=row, column=6, value=f'=IFERROR((D{row}-C{row})/C{row},"")')
        for col in range(1, 7):
            fmt_cell(ws.cell(row=row, column=col))
        ws.cell(row=row, column=3).number_format = CURRENCY
        ws.cell(row=row, column=4).number_format = CURRENCY
        ws.cell(row=row, column=5).number_format = CURRENCY
        ws.cell(row=row, column=6).number_format = PCT_FMT
        row += 1
    if drop:
        ws.conditional_formatting.add(f'F{start}:F{row-1}',
            CellIsRule(operator='lessThan', formula=['0'], fill=PatternFill('solid', start_color=COLOR_DROP)))
    else:
        ws.conditional_formatting.add(f'F{start}:F{row-1}',
            CellIsRule(operator='greaterThan', formula=['0'], fill=PatternFill('solid', start_color=COLOR_GROW)))
    return row + 2

row = build_trend_block(ws, row, '📉 PRODUCTOS EN DECLIVE — top 10 mayor caída (3M reciente vs 3M previo, base ≥ $500k 6M)', '9C0006', declive, drop=True)
row = build_trend_block(ws, row, '📈 PRODUCTOS EN ASCENSO — top 10 mayor crecimiento (3M reciente vs 3M previo, base ≥ $500k 6M)', '006100', ascenso, drop=False)

# Cola larga: productos que aparecen en facturas con frecuencia < 1/mes
meses_periodo = (MAX_DATE.to_period('M').ordinal - df['Fecha'].min().to_period('M').ordinal) + 1
prod_freq = df.groupby('Cod Interno').agg(
    Facturas=('Factura', 'nunique'),
    Ingresos=('Total', 'sum'),
    Unidades=('Cantidad', 'sum'),
    Desc=('Descripcion Producto', 'first'),
    Cat=('CategoriaProducto', 'first'),
    FechaMax=('Fecha', 'max')
)
prod_freq['FacturasPorMes'] = prod_freq['Facturas'] / meses_periodo
cola_larga = prod_freq[prod_freq['FacturasPorMes'] < 1].sort_values('Ingresos').head(50)

ws.cell(row=row, column=1, value=f'🐌 PRODUCTOS COLA LARGA — frecuencia < 1 factura/mes (top 50 menos ingresos)').font = Font(name=FONT_NAME, size=12, bold=True, color='595959')
row += 1
ws.cell(row=row, column=1, value=f'Total productos cola larga: {len(prod_freq[prod_freq["FacturasPorMes"] < 1])} de {len(prod_freq)} SKUs').font = Font(name=FONT_NAME, size=10, italic=True)
row += 1
headers = ['Producto', 'Categoría', 'Ingresos', 'Unidades', 'Nº Facturas', 'Fact/mes', 'Última venta']
for i, h in enumerate(headers):
    c = ws.cell(row=row, column=i+1, value=h); fmt_header(c)
row += 1
for _, r in cola_larga.iterrows():
    ws.cell(row=row, column=1, value=r['Desc'])
    ws.cell(row=row, column=2, value=r['Cat'])
    ws.cell(row=row, column=3, value=int(r['Ingresos']))
    ws.cell(row=row, column=4, value=int(r['Unidades']))
    ws.cell(row=row, column=5, value=int(r['Facturas']))
    ws.cell(row=row, column=6, value=round(r['FacturasPorMes'], 2))
    ws.cell(row=row, column=7, value=r['FechaMax'].date())
    for col in range(1, 8):
        fmt_cell(ws.cell(row=row, column=col))
    ws.cell(row=row, column=3).number_format = CURRENCY
    ws.cell(row=row, column=4).number_format = INT_FMT
    ws.cell(row=row, column=5).number_format = INT_FMT
    ws.cell(row=row, column=6).number_format = '0.00'
    ws.cell(row=row, column=7).number_format = 'yyyy-mm-dd'
    row += 1

autosize(ws, {'A': 55, 'B': 24, 'C': 16, 'D': 18, 'E': 18, 'F': 14, 'G': 14})

# =========================================================
# HOJA 8: CONCENTRACIÓN Y RIESGO
# =========================================================
print('Construyendo HOJA 8: Concentración y Riesgo...')
ws = wb.create_sheet('8. Concentracion Riesgo')
ws['A1'] = 'CONCENTRACIÓN Y RIESGO'
fmt_title(ws['A1'])
ws.merge_cells('A1:E1')

# Concentración clientes
cli_sorted = cli_real.sort_values('Ingresos', ascending=False).reset_index(drop=True)
total_cli_real = cli_sorted['Ingresos'].sum()
top5_cli = cli_sorted.head(5)['Ingresos'].sum()
top10_cli = cli_sorted.head(10)['Ingresos'].sum()
top20_cli = cli_sorted.head(20)['Ingresos'].sum()

# Concentración productos
prod_sorted = prod_agg.sort_values('Ingresos', ascending=False).reset_index(drop=True)
top5_prod = prod_sorted.head(5)['Ingresos'].sum()
top10_prod = prod_sorted.head(10)['Ingresos'].sum()
top20_prod = prod_sorted.head(20)['Ingresos'].sum()

row = 3
ws.cell(row=row, column=1, value='CONCENTRACIÓN DE CLIENTES (excluye CONSUMIDOR FINAL)').font = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_HEADER)
row += 1
ws.cell(row=row, column=1, value=f'Total clientes reales: {len(cli_sorted)}    |    Facturación real: ${total_cli_real:,.0f}').font = Font(name=FONT_NAME, size=10, italic=True)
row += 2
headers = ['Grupo', 'Facturación', '% del total real']
for i, h in enumerate(headers):
    c = ws.cell(row=row, column=i+1, value=h); fmt_header(c)
row += 1
for label, v in [('Top 5 clientes', top5_cli), ('Top 10 clientes', top10_cli), ('Top 20 clientes', top20_cli)]:
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=int(v))
    ws.cell(row=row, column=3, value=f'=B{row}/{int(total_cli_real)}')
    for col in range(1, 4):
        fmt_cell(ws.cell(row=row, column=col))
    ws.cell(row=row, column=2).number_format = CURRENCY
    ws.cell(row=row, column=3).number_format = PCT_FMT
    row += 1

row += 2
ws.cell(row=row, column=1, value='CONCENTRACIÓN DE PRODUCTOS').font = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_HEADER)
row += 1
ws.cell(row=row, column=1, value=f'Total SKUs: {len(prod_sorted)}    |    Facturación: ${total_global:,.0f}').font = Font(name=FONT_NAME, size=10, italic=True)
row += 2
for i, h in enumerate(headers):
    c = ws.cell(row=row, column=i+1, value=h); fmt_header(c)
row += 1
for label, v in [('Top 5 productos', top5_prod), ('Top 10 productos', top10_prod), ('Top 20 productos', top20_prod)]:
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=int(v))
    ws.cell(row=row, column=3, value=f'=B{row}/{int(total_global)}')
    for col in range(1, 4):
        fmt_cell(ws.cell(row=row, column=col))
    ws.cell(row=row, column=2).number_format = CURRENCY
    ws.cell(row=row, column=3).number_format = PCT_FMT
    row += 1

# Análisis Pareto
row += 2
ws.cell(row=row, column=1, value='ANÁLISIS PARETO (80/20)').font = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_HEADER)
row += 1
# Productos para llegar al 80%
prod_sorted['Acum'] = prod_sorted['Ingresos'].cumsum()
prod_sorted['AcumPct'] = prod_sorted['Acum'] / total_global
n_prod_80 = (prod_sorted['AcumPct'] <= 0.80).sum() + 1
pct_skus_80 = n_prod_80 / len(prod_sorted)

cli_sorted['Acum'] = cli_sorted['Ingresos'].cumsum()
cli_sorted['AcumPct'] = cli_sorted['Acum'] / total_cli_real
n_cli_80 = (cli_sorted['AcumPct'] <= 0.80).sum() + 1
pct_cli_80 = n_cli_80 / len(cli_sorted)

headers = ['Métrica', 'Cantidad', '% del total', 'Genera']
for i, h in enumerate(headers):
    c = ws.cell(row=row, column=i+1, value=h); fmt_header(c)
row += 1

ws.cell(row=row, column=1, value='Productos que generan el 80% de la facturación')
ws.cell(row=row, column=2, value=n_prod_80)
ws.cell(row=row, column=3, value=pct_skus_80)
ws.cell(row=row, column=4, value='≈ 80%')
for col in range(1, 5):
    fmt_cell(ws.cell(row=row, column=col))
ws.cell(row=row, column=2).number_format = INT_FMT
ws.cell(row=row, column=3).number_format = PCT_FMT
row += 1

ws.cell(row=row, column=1, value='Clientes que generan el 80% de la facturación (excl. CF)')
ws.cell(row=row, column=2, value=n_cli_80)
ws.cell(row=row, column=3, value=pct_cli_80)
ws.cell(row=row, column=4, value='≈ 80%')
for col in range(1, 5):
    fmt_cell(ws.cell(row=row, column=col))
ws.cell(row=row, column=2).number_format = INT_FMT
ws.cell(row=row, column=3).number_format = PCT_FMT
row += 1

autosize(ws, {'A': 55, 'B': 18, 'C': 16, 'D': 14, 'E': 14})

# =========================================================
# HOJA 9: OPORTUNIDADES COMERCIALES
# =========================================================
print('Construyendo HOJA 9: Oportunidades Comerciales...')
ws = wb.create_sheet('9. Oportunidades')
ws['A1'] = 'OPORTUNIDADES COMERCIALES'
fmt_title(ws['A1'])
ws.merge_cells('A1:F1')

# 1) Cross-selling: pares de productos en misma factura
print('  Calculando cross-selling (puede tardar)...')
# Para optimizar, agrupamos por factura los códigos. Tomamos solo los top 200 productos para reducir combinaciones
top_prods = set(prod_agg.nlargest(200, 'Ingresos')['Cod Interno'])
fact_groups = df[df['Cod Interno'].isin(top_prods)].groupby('Factura')['Cod Interno'].apply(set)
fact_groups = fact_groups[fact_groups.apply(len) >= 2]

pair_counter = Counter()
for prods_set in fact_groups:
    if len(prods_set) > 30:  # facturas muy grandes generan muchas combinaciones, las limitamos
        continue
    for pair in combinations(sorted(prods_set), 2):
        pair_counter[pair] += 1

top_pairs = pair_counter.most_common(15)
prod_name_map = prod_agg.set_index('Cod Interno')['Descripcion Producto'].to_dict()
prod_cat_map = prod_agg.set_index('Cod Interno')['CategoriaProducto'].to_dict()

row = 3
ws.cell(row=row, column=1, value='CROSS-SELLING — TOP 15 pares de productos comprados juntos').font = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_HEADER)
row += 1
ws.cell(row=row, column=1, value=f'Calculado sobre top 200 productos. Total facturas con ≥2 items: {len(fact_groups)}').font = Font(name=FONT_NAME, size=9, italic=True)
row += 2
headers = ['#', 'Producto A', 'Categoría A', 'Producto B', 'Categoría B', 'Nº Facturas en común']
for i, h in enumerate(headers):
    c = ws.cell(row=row, column=i+1, value=h); fmt_header(c)
row += 1
for i, (pair, cnt) in enumerate(top_pairs):
    a, b = pair
    ws.cell(row=row, column=1, value=i+1)
    ws.cell(row=row, column=2, value=prod_name_map.get(a, str(a)))
    ws.cell(row=row, column=3, value=prod_cat_map.get(a, ''))
    ws.cell(row=row, column=4, value=prod_name_map.get(b, str(b)))
    ws.cell(row=row, column=5, value=prod_cat_map.get(b, ''))
    ws.cell(row=row, column=6, value=cnt)
    for col in range(1, 7):
        fmt_cell(ws.cell(row=row, column=col))
    ws.cell(row=row, column=6).number_format = INT_FMT
    row += 1

# 2) Clientes con poca diversidad de categorías
row += 2
ws.cell(row=row, column=1, value='CLIENTES POCO DIVERSIFICADOS — Top 20 facturación que compran 1-2 categorías (oportunidad mix)').font = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_HEADER)
row += 1
cat_por_cli = df.groupby('Cliente')['CategoriaProducto'].nunique().reset_index()
cat_por_cli.columns = ['Cliente', 'NumCategorias']
cli_div = cli_real.merge(cat_por_cli, on='Cliente')
poco_div = cli_div[cli_div['NumCategorias'] <= 2].sort_values('Ingresos', ascending=False).head(20)

headers = ['#', 'Cliente', 'Facturación', 'Nº Facturas', 'Nº Categorías', 'Categorías que compra']
for i, h in enumerate(headers):
    c = ws.cell(row=row, column=i+1, value=h); fmt_header(c)
row += 1
cli_cats = df.groupby('Cliente')['CategoriaProducto'].apply(lambda x: ' | '.join(sorted(x.unique()))).to_dict()
for i, r in poco_div.iterrows():
    ws.cell(row=row, column=1, value=i+1 if False else len([c for c in poco_div.iloc[:list(poco_div.index).index(i)+1].index]) )
ws_cell_idx = 0
# rebuilding loop cleaner:
poco_div = poco_div.reset_index(drop=True)
# fix: we already wrote partial row, let's clear and rebuild
# Actually, the bad cell above was in current row. Let's just overwrite by going back.
# Simpler: re-rewrite. Decrement row variable wasn't bumped yet so still at correct row.
# Let's overwrite the cell we erroneously wrote
ws.cell(row=row, column=1, value='')  # clear
for i, r in poco_div.iterrows():
    ws.cell(row=row, column=1, value=int(i)+1)
    ws.cell(row=row, column=2, value=r['Cliente'])
    ws.cell(row=row, column=3, value=int(r['Ingresos']))
    ws.cell(row=row, column=4, value=int(r['Facturas']))
    ws.cell(row=row, column=5, value=int(r['NumCategorias']))
    ws.cell(row=row, column=6, value=cli_cats.get(r['Cliente'], ''))
    for col in range(1, 7):
        fmt_cell(ws.cell(row=row, column=col))
    ws.cell(row=row, column=3).number_format = CURRENCY
    ws.cell(row=row, column=4).number_format = INT_FMT
    ws.cell(row=row, column=5).number_format = INT_FMT
    row += 1

# 3) Productos que compran clientes top y no compran clientes medianos
row += 2
ws.cell(row=row, column=1, value='RECOMENDACIÓN — productos populares en clientes A que clientes B casi no compran').font = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_HEADER)
row += 1

# Segmento A (top 20%) y B (siguiente 30%)
cli_sorted_local = cli_real.sort_values('Ingresos', ascending=False).reset_index(drop=True)
n_total_cli = len(cli_sorted_local)
n_a_local = max(1, int(round(n_total_cli * 0.2)))
n_b_local = max(1, int(round(n_total_cli * 0.3)))
seg_a_cli = set(cli_sorted_local.iloc[:n_a_local]['Cliente'])
seg_b_cli = set(cli_sorted_local.iloc[n_a_local:n_a_local+n_b_local]['Cliente'])

prod_in_a = df[df['Cliente'].isin(seg_a_cli)].groupby('Cod Interno').agg(
    ClientesA=('Cliente', 'nunique'),
    IngresosA=('Total', 'sum'),
    Desc=('Descripcion Producto', 'first'),
    Cat=('CategoriaProducto', 'first')
)
prod_in_b = df[df['Cliente'].isin(seg_b_cli)].groupby('Cod Interno').agg(
    ClientesB=('Cliente', 'nunique')
)
prod_recomend = prod_in_a.join(prod_in_b, how='left').fillna(0)
prod_recomend['PenA'] = prod_recomend['ClientesA'] / len(seg_a_cli)
prod_recomend['PenB'] = prod_recomend['ClientesB'] / len(seg_b_cli)
prod_recomend['GapPen'] = prod_recomend['PenA'] - prod_recomend['PenB']
# Buscamos productos populares en A (penetración > 30%) y bajos en B (penetración < 10%)
oport = prod_recomend[(prod_recomend['PenA'] >= 0.30) & (prod_recomend['PenB'] < 0.10)].sort_values('GapPen', ascending=False).head(15)

headers = ['Producto', 'Categoría', '% clientes A que lo compran', '% clientes B que lo compran', 'Gap', 'Ingresos clientes A']
for i, h in enumerate(headers):
    c = ws.cell(row=row, column=i+1, value=h); fmt_header(c)
row += 1
for cod, r in oport.iterrows():
    ws.cell(row=row, column=1, value=r['Desc'])
    ws.cell(row=row, column=2, value=r['Cat'])
    ws.cell(row=row, column=3, value=r['PenA'])
    ws.cell(row=row, column=4, value=r['PenB'])
    ws.cell(row=row, column=5, value=r['GapPen'])
    ws.cell(row=row, column=6, value=int(r['IngresosA']))
    for col in range(1, 7):
        fmt_cell(ws.cell(row=row, column=col))
    ws.cell(row=row, column=3).number_format = PCT_FMT
    ws.cell(row=row, column=4).number_format = PCT_FMT
    ws.cell(row=row, column=5).number_format = PCT_FMT
    ws.cell(row=row, column=6).number_format = CURRENCY
    row += 1

autosize(ws, {'A': 50, 'B': 24, 'C': 18, 'D': 50, 'E': 22, 'F': 22})

wb.save('Analisis_Global_Family_Mayo2026.xlsx')
print('Hojas 7-9 OK, archivo guardado.')
