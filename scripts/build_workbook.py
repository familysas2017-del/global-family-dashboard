"""
Construye el Excel final: Analisis_Global_Family_Mayo2026.xlsx
Lee df_clean.pkl generado por build_analysis.py
"""
import pandas as pd
import numpy as np
import warnings
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, FormulaRule
from openpyxl.utils import get_column_letter
from itertools import combinations
from collections import Counter

warnings.filterwarnings('ignore')

df = pd.read_pickle('df_clean.pkl')
print(f'Cargados {len(df)} registros')

# Fechas de referencia
MAX_DATE = df['Fecha'].max()
MIN_DATE = df['Fecha'].min()
LAST_3M_START = MAX_DATE - pd.Timedelta(days=90)
PREV_3M_START = MAX_DATE - pd.Timedelta(days=180)
LAST_30D = MAX_DATE - pd.Timedelta(days=30)
LAST_60D = MAX_DATE - pd.Timedelta(days=60)
LAST_90D = MAX_DATE - pd.Timedelta(days=90)

print(f'Fecha máx: {MAX_DATE}, fecha mín: {MIN_DATE}')

# =========================================================
# ESTILOS
# =========================================================
FONT_NAME = 'Arial'

# Colores corporativos
COLOR_HEADER = '1F4E78'   # azul oscuro
COLOR_SUBHEADER = '5B9BD5' # azul medio
COLOR_TOTAL = 'D9E1F2'    # azul claro
COLOR_ALERT_BG = 'F8CBAD'  # naranja claro
COLOR_GROW = 'C6EFCE'     # verde claro
COLOR_DROP = 'FFC7CE'     # rojo claro

thin = Side(border_style='thin', color='BFBFBF')
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

def fmt_header(cell):
    cell.font = Font(name=FONT_NAME, size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', start_color=COLOR_HEADER)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_all

def fmt_subheader(cell):
    cell.font = Font(name=FONT_NAME, size=10, bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', start_color=COLOR_SUBHEADER)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_all

def fmt_title(cell):
    cell.font = Font(name=FONT_NAME, size=14, bold=True, color=COLOR_HEADER)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def fmt_total_row(cell):
    cell.font = Font(name=FONT_NAME, size=10, bold=True)
    cell.fill = PatternFill('solid', start_color=COLOR_TOTAL)
    cell.border = border_all

def fmt_cell(cell, num_fmt=None, bold=False):
    cell.font = Font(name=FONT_NAME, size=10, bold=bold)
    cell.border = border_all
    if num_fmt:
        cell.number_format = num_fmt

def autosize(ws, widths):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

CURRENCY = '$#,##0;($#,##0);-'
INT_FMT = '#,##0;(#,##0);-'
PCT_FMT = '0.0%;(0.0%);-'
DATE_FMT = 'yyyy-mm-dd'

# =========================================================
# CREAR WORKBOOK
# =========================================================
wb = Workbook()
wb.remove(wb.active)  # quitar default

# =========================================================
# HOJA 1: RESUMEN EJECUTIVO
# =========================================================
print('Construyendo HOJA 1: Resumen Ejecutivo...')
ws = wb.create_sheet('1. Resumen Ejecutivo')

# Métricas base (calculadas en pandas, escritas como valores)
fact_total = df['Total'].sum()
n_facturas = df['Factura'].nunique()
n_lineas = len(df)
ticket_prom = fact_total / n_facturas
n_clientes = df['Cliente'].nunique()
n_clientes_no_gen = df[~df['ClienteGenerico']]['Cliente'].nunique()
n_skus = df['Cod Interno'].nunique()
unidades_tot = df['Cantidad'].sum()

# Facturación por mes
mes_fact = df.groupby('AñoMes')['Total'].sum().sort_index()
mes_max = mes_fact.idxmax()
mes_min = mes_fact.idxmin()

# Construir hoja
ws['A1'] = 'RESUMEN EJECUTIVO — GLOBAL FAMILY DISTRIBUCIONES'
fmt_title(ws['A1'])
ws.merge_cells('A1:F1')

ws['A2'] = f'Período analizado: {MIN_DATE.strftime("%Y-%m-%d")} a {MAX_DATE.strftime("%Y-%m-%d")}'
ws['A2'].font = Font(name=FONT_NAME, size=10, italic=True, color='595959')

# KPIs principales
ws['A4'] = 'KPI'
ws['B4'] = 'Valor'
fmt_header(ws['A4']); fmt_header(ws['B4'])

kpis = [
    ('Facturación total histórica', fact_total, CURRENCY),
    ('Número total de facturas únicas', n_facturas, INT_FMT),
    ('Número total de líneas de factura', n_lineas, INT_FMT),
    ('Ticket promedio por factura', ticket_prom, CURRENCY),
    ('Unidades totales vendidas', unidades_tot, INT_FMT),
    ('Clientes únicos (todos)', n_clientes, INT_FMT),
    ('Clientes únicos (excl. Consumidor Final)', n_clientes_no_gen, INT_FMT),
    ('Productos / SKUs únicos vendidos', n_skus, INT_FMT),
    ('Mes con MAYOR facturación', f'{mes_max}   →   ${mes_fact.max():,.0f}', None),
    ('Mes con MENOR facturación', f'{mes_min}   →   ${mes_fact.min():,.0f}', None),
]
row = 5
for k, v, fmt in kpis:
    ws.cell(row=row, column=1, value=k)
    c = ws.cell(row=row, column=2, value=v)
    fmt_cell(ws.cell(row=row, column=1), bold=True)
    fmt_cell(c, num_fmt=fmt)
    row += 1

# Facturación por mes (tabla)
row += 2
ws.cell(row=row, column=1, value='FACTURACIÓN POR MES').font = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_HEADER)
row += 1
headers = ['Año-Mes', 'Facturación', 'Nº Facturas', 'Ticket Promedio', 'Unidades', 'Crec. m/m %']
for i, h in enumerate(headers):
    c = ws.cell(row=row, column=i+1, value=h); fmt_header(c)
row += 1
start_row = row

mes_data = df.groupby('AñoMes').agg(
    Total=('Total', 'sum'),
    Facturas=('Factura', 'nunique'),
    Unidades=('Cantidad', 'sum')
).sort_index()

prev_total = None
for am, r in mes_data.iterrows():
    ws.cell(row=row, column=1, value=am)
    ws.cell(row=row, column=2, value=int(r['Total']))
    ws.cell(row=row, column=3, value=int(r['Facturas']))
    # Ticket = Total/Facturas como fórmula
    ws.cell(row=row, column=4, value=f'=B{row}/C{row}')
    ws.cell(row=row, column=5, value=int(r['Unidades']))
    # Crecimiento m/m fórmula
    if row > start_row:
        ws.cell(row=row, column=6, value=f'=IFERROR((B{row}-B{row-1})/B{row-1},"")')
    else:
        ws.cell(row=row, column=6, value='—')
    for col in range(1, 7):
        fmt_cell(ws.cell(row=row, column=col))
    ws.cell(row=row, column=2).number_format = CURRENCY
    ws.cell(row=row, column=3).number_format = INT_FMT
    ws.cell(row=row, column=4).number_format = CURRENCY
    ws.cell(row=row, column=5).number_format = INT_FMT
    ws.cell(row=row, column=6).number_format = PCT_FMT
    row += 1

# Fila total
ws.cell(row=row, column=1, value='TOTAL')
ws.cell(row=row, column=2, value=f'=SUM(B{start_row}:B{row-1})')
ws.cell(row=row, column=3, value=f'=SUM(C{start_row}:C{row-1})')
ws.cell(row=row, column=4, value=f'=IFERROR(B{row}/C{row},0)')
ws.cell(row=row, column=5, value=f'=SUM(E{start_row}:E{row-1})')
ws.cell(row=row, column=6, value='—')
for col in range(1, 7):
    fmt_total_row(ws.cell(row=row, column=col))
ws.cell(row=row, column=2).number_format = CURRENCY
ws.cell(row=row, column=4).number_format = CURRENCY
ws.cell(row=row, column=5).number_format = INT_FMT

# Formato condicional para crec. m/m
last_data_row = row - 1
ws.conditional_formatting.add(
    f'F{start_row+1}:F{last_data_row}',
    CellIsRule(operator='greaterThan', formula=['0'], fill=PatternFill('solid', start_color=COLOR_GROW))
)
ws.conditional_formatting.add(
    f'F{start_row+1}:F{last_data_row}',
    CellIsRule(operator='lessThan', formula=['0'], fill=PatternFill('solid', start_color=COLOR_DROP))
)

# Nota sobre YoY
row += 2
ws.cell(row=row, column=1, value='Nota: el histórico cubre solo Oct 2025 – May 2026 (~7 meses), por lo que NO se incluyen comparativos año vs año.')
ws.cell(row=row, column=1).font = Font(name=FONT_NAME, size=9, italic=True, color='C00000')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

autosize(ws, {'A': 42, 'B': 22, 'C': 14, 'D': 18, 'E': 14, 'F': 14})

print(f'  Hoja 1 OK: {row} filas')

# =========================================================
# HOJA 2: CATEGORÍAS (PRODUCTO)
# =========================================================
print('Construyendo HOJA 2: Categorías (Producto)...')
ws = wb.create_sheet('2. Categorias Producto')

ws['A1'] = 'CATEGORÍAS DE PRODUCTO (inferidas desde descripción)'
fmt_title(ws['A1'])
ws.merge_cells('A1:I1')

ws['A2'] = 'Las categorías se infieren con reglas de palabras clave aplicadas a la descripción del producto. Ver "Notas Metodológicas".'
ws['A2'].font = Font(name=FONT_NAME, size=9, italic=True, color='595959')
ws.merge_cells('A2:I2')

headers = ['Categoría', 'Ingresos', '% Total', 'Unidades', 'SKUs', 'Ticket Prom (línea)', 'Ventas 3M recientes', 'Ventas 3M previos', 'Δ% 3M']
for i, h in enumerate(headers):
    c = ws.cell(row=4, column=i+1, value=h); fmt_header(c)

cat_data = df.groupby('CategoriaProducto').agg(
    Ingresos=('Total', 'sum'),
    Unidades=('Cantidad', 'sum'),
    SKUs=('Cod Interno', 'nunique'),
    Lineas=('Total', 'count')
).sort_values('Ingresos', ascending=False)

# Tendencia: 3M recientes vs 3M previos
df_3m = df[df['Fecha'] >= LAST_3M_START]
df_3m_prev = df[(df['Fecha'] >= PREV_3M_START) & (df['Fecha'] < LAST_3M_START)]
cat_3m = df_3m.groupby('CategoriaProducto')['Total'].sum()
cat_3m_prev = df_3m_prev.groupby('CategoriaProducto')['Total'].sum()

total_global = df['Total'].sum()
row = 5
start_row = row
for cat, r in cat_data.iterrows():
    v3m = cat_3m.get(cat, 0)
    v3mp = cat_3m_prev.get(cat, 0)
    ws.cell(row=row, column=1, value=cat)
    ws.cell(row=row, column=2, value=int(r['Ingresos']))
    ws.cell(row=row, column=3, value=f'=B{row}/${chr(66)}${start_row + len(cat_data)}')  # % vs total
    ws.cell(row=row, column=4, value=int(r['Unidades']))
    ws.cell(row=row, column=5, value=int(r['SKUs']))
    ws.cell(row=row, column=6, value=f'=IFERROR(B{row}/{int(r["Lineas"])},0)')
    ws.cell(row=row, column=7, value=int(v3m))
    ws.cell(row=row, column=8, value=int(v3mp))
    ws.cell(row=row, column=9, value=f'=IFERROR((G{row}-H{row})/H{row},"")')
    for col in range(1, 10):
        fmt_cell(ws.cell(row=row, column=col))
    ws.cell(row=row, column=2).number_format = CURRENCY
    ws.cell(row=row, column=3).number_format = PCT_FMT
    ws.cell(row=row, column=4).number_format = INT_FMT
    ws.cell(row=row, column=5).number_format = INT_FMT
    ws.cell(row=row, column=6).number_format = CURRENCY
    ws.cell(row=row, column=7).number_format = CURRENCY
    ws.cell(row=row, column=8).number_format = CURRENCY
    ws.cell(row=row, column=9).number_format = PCT_FMT
    row += 1

# Total
ws.cell(row=row, column=1, value='TOTAL')
ws.cell(row=row, column=2, value=f'=SUM(B{start_row}:B{row-1})')
ws.cell(row=row, column=3, value=f'=SUM(C{start_row}:C{row-1})')
ws.cell(row=row, column=4, value=f'=SUM(D{start_row}:D{row-1})')
ws.cell(row=row, column=5, value='')
ws.cell(row=row, column=6, value='')
ws.cell(row=row, column=7, value=f'=SUM(G{start_row}:G{row-1})')
ws.cell(row=row, column=8, value=f'=SUM(H{start_row}:H{row-1})')
ws.cell(row=row, column=9, value=f'=IFERROR((G{row}-H{row})/H{row},"")')
for col in range(1, 10):
    fmt_total_row(ws.cell(row=row, column=col))
ws.cell(row=row, column=2).number_format = CURRENCY
ws.cell(row=row, column=3).number_format = PCT_FMT
ws.cell(row=row, column=4).number_format = INT_FMT
ws.cell(row=row, column=7).number_format = CURRENCY
ws.cell(row=row, column=8).number_format = CURRENCY
ws.cell(row=row, column=9).number_format = PCT_FMT

# Formato condicional Δ%
ws.conditional_formatting.add(
    f'I{start_row}:I{row-1}',
    CellIsRule(operator='greaterThan', formula=['0'], fill=PatternFill('solid', start_color=COLOR_GROW))
)
ws.conditional_formatting.add(
    f'I{start_row}:I{row-1}',
    CellIsRule(operator='lessThan', formula=['0'], fill=PatternFill('solid', start_color=COLOR_DROP))
)

autosize(ws, {'A': 35, 'B': 18, 'C': 10, 'D': 14, 'E': 8, 'F': 16, 'G': 18, 'H': 18, 'I': 10})

# =========================================================
# HOJA 2b: CATEGORÍAS POR PROVEEDOR
# =========================================================
print('Construyendo HOJA 2b: Categorías por Proveedor...')
ws = wb.create_sheet('2b. Categorias Proveedor')
ws['A1'] = 'CATEGORÍAS POR PROVEEDOR / MARCA (valor original del archivo)'
fmt_title(ws['A1'])
ws.merge_cells('A1:I1')

headers = ['Proveedor / Marca', 'Ingresos', '% Total', 'Unidades', 'SKUs', 'Ticket Prom (línea)', 'Ventas 3M recientes', 'Ventas 3M previos', 'Δ% 3M']
for i, h in enumerate(headers):
    c = ws.cell(row=4, column=i+1, value=h); fmt_header(c)

prov_data = df.groupby('Proveedor').agg(
    Ingresos=('Total', 'sum'),
    Unidades=('Cantidad', 'sum'),
    SKUs=('Cod Interno', 'nunique'),
    Lineas=('Total', 'count')
).sort_values('Ingresos', ascending=False)

prov_3m = df_3m.groupby('Proveedor')['Total'].sum()
prov_3m_prev = df_3m_prev.groupby('Proveedor')['Total'].sum()

row = 5
start_row = row
for prov, r in prov_data.iterrows():
    v3m = prov_3m.get(prov, 0)
    v3mp = prov_3m_prev.get(prov, 0)
    ws.cell(row=row, column=1, value=prov)
    ws.cell(row=row, column=2, value=int(r['Ingresos']))
    ws.cell(row=row, column=3, value=f'=B{row}/$B${start_row + len(prov_data)}')
    ws.cell(row=row, column=4, value=int(r['Unidades']))
    ws.cell(row=row, column=5, value=int(r['SKUs']))
    ws.cell(row=row, column=6, value=f'=IFERROR(B{row}/{int(r["Lineas"])},0)')
    ws.cell(row=row, column=7, value=int(v3m))
    ws.cell(row=row, column=8, value=int(v3mp))
    ws.cell(row=row, column=9, value=f'=IFERROR((G{row}-H{row})/H{row},"")')
    for col in range(1, 10):
        fmt_cell(ws.cell(row=row, column=col))
    ws.cell(row=row, column=2).number_format = CURRENCY
    ws.cell(row=row, column=3).number_format = PCT_FMT
    ws.cell(row=row, column=4).number_format = INT_FMT
    ws.cell(row=row, column=5).number_format = INT_FMT
    ws.cell(row=row, column=6).number_format = CURRENCY
    ws.cell(row=row, column=7).number_format = CURRENCY
    ws.cell(row=row, column=8).number_format = CURRENCY
    ws.cell(row=row, column=9).number_format = PCT_FMT
    row += 1

ws.cell(row=row, column=1, value='TOTAL')
ws.cell(row=row, column=2, value=f'=SUM(B{start_row}:B{row-1})')
ws.cell(row=row, column=3, value=f'=SUM(C{start_row}:C{row-1})')
ws.cell(row=row, column=4, value=f'=SUM(D{start_row}:D{row-1})')
ws.cell(row=row, column=7, value=f'=SUM(G{start_row}:G{row-1})')
ws.cell(row=row, column=8, value=f'=SUM(H{start_row}:H{row-1})')
ws.cell(row=row, column=9, value=f'=IFERROR((G{row}-H{row})/H{row},"")')
for col in range(1, 10):
    fmt_total_row(ws.cell(row=row, column=col))
ws.cell(row=row, column=2).number_format = CURRENCY
ws.cell(row=row, column=3).number_format = PCT_FMT
ws.cell(row=row, column=4).number_format = INT_FMT
ws.cell(row=row, column=7).number_format = CURRENCY
ws.cell(row=row, column=8).number_format = CURRENCY
ws.cell(row=row, column=9).number_format = PCT_FMT

ws.conditional_formatting.add(
    f'I{start_row}:I{row-1}',
    CellIsRule(operator='greaterThan', formula=['0'], fill=PatternFill('solid', start_color=COLOR_GROW))
)
ws.conditional_formatting.add(
    f'I{start_row}:I{row-1}',
    CellIsRule(operator='lessThan', formula=['0'], fill=PatternFill('solid', start_color=COLOR_DROP))
)

autosize(ws, {'A': 28, 'B': 18, 'C': 10, 'D': 14, 'E': 8, 'F': 16, 'G': 18, 'H': 18, 'I': 10})

# =========================================================
# HOJA 3: TOP 20 PRODUCTOS POR FACTURACIÓN
# =========================================================
print('Construyendo HOJA 3: Top 20 productos por facturación...')
ws = wb.create_sheet('3. Top20 por Facturacion')

ws['A1'] = 'TOP 20 PRODUCTOS POR FACTURACIÓN'
fmt_title(ws['A1'])
ws.merge_cells('A1:I1')

prod_agg = df.groupby(['Cod Interno', 'Descripcion Producto', 'CategoriaProducto', 'Proveedor']).agg(
    Ingresos=('Total', 'sum'),
    Unidades=('Cantidad', 'sum'),
    Lineas=('Factura', 'count'),
    Facturas=('Factura', 'nunique'),
    Clientes=('Cliente', 'nunique'),
    FechaMin=('Fecha', 'min'),
    FechaMax=('Fecha', 'max')
).reset_index()
prod_agg['PrecioProm'] = prod_agg['Ingresos'] / prod_agg['Unidades']
# Frecuencia: días entre fechaMin y fechaMax / nº facturas
prod_agg['DiasActivo'] = (prod_agg['FechaMax'] - prod_agg['FechaMin']).dt.days + 1
prod_agg['FreqDias'] = prod_agg['DiasActivo'] / prod_agg['Facturas']

top20_fact = prod_agg.nlargest(20, 'Ingresos').reset_index(drop=True)

headers = ['#', 'Producto', 'Categoría', 'Proveedor', 'Ingresos', '% Total', 'Unidades', 'Precio Prom.', 'Nº Clientes', 'Frec. compra (días)']
for i, h in enumerate(headers):
    c = ws.cell(row=3, column=i+1, value=h); fmt_header(c)

row = 4
start_row = row
for i, r in top20_fact.iterrows():
    ws.cell(row=row, column=1, value=i+1)
    ws.cell(row=row, column=2, value=r['Descripcion Producto'])
    ws.cell(row=row, column=3, value=r['CategoriaProducto'])
    ws.cell(row=row, column=4, value=r['Proveedor'])
    ws.cell(row=row, column=5, value=int(r['Ingresos']))
    ws.cell(row=row, column=6, value=f'=E{row}/{int(total_global)}')
    ws.cell(row=row, column=7, value=int(r['Unidades']))
    ws.cell(row=row, column=8, value=f'=IFERROR(E{row}/G{row},0)')
    ws.cell(row=row, column=9, value=int(r['Clientes']))
    ws.cell(row=row, column=10, value=round(r['FreqDias'], 1))
    for col in range(1, 11):
        fmt_cell(ws.cell(row=row, column=col))
    ws.cell(row=row, column=5).number_format = CURRENCY
    ws.cell(row=row, column=6).number_format = PCT_FMT
    ws.cell(row=row, column=7).number_format = INT_FMT
    ws.cell(row=row, column=8).number_format = CURRENCY
    ws.cell(row=row, column=9).number_format = INT_FMT
    ws.cell(row=row, column=10).number_format = '0.0'
    row += 1

# Total top 20
ws.cell(row=row, column=1, value='')
ws.cell(row=row, column=2, value='TOTAL TOP 20')
ws.cell(row=row, column=5, value=f'=SUM(E{start_row}:E{row-1})')
ws.cell(row=row, column=6, value=f'=SUM(F{start_row}:F{row-1})')
ws.cell(row=row, column=7, value=f'=SUM(G{start_row}:G{row-1})')
ws.cell(row=row, column=9, value=f'=SUM(I{start_row}:I{row-1})')
for col in range(1, 11):
    fmt_total_row(ws.cell(row=row, column=col))
ws.cell(row=row, column=5).number_format = CURRENCY
ws.cell(row=row, column=6).number_format = PCT_FMT
ws.cell(row=row, column=7).number_format = INT_FMT
ws.cell(row=row, column=9).number_format = INT_FMT

autosize(ws, {'A': 4, 'B': 55, 'C': 24, 'D': 16, 'E': 18, 'F': 9, 'G': 12, 'H': 16, 'I': 12, 'J': 16})

# Guardar prod_agg para siguientes hojas
prod_agg.to_pickle('prod_agg.pkl')
top20_fact.to_pickle('top20_fact.pkl')

wb.save('Analisis_Global_Family_Mayo2026.xlsx')
print('Workbook guardado (hojas 1-3).')
