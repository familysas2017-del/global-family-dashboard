"""
HOJA 10: Dashboard Datos
+ Notas Metodológicas
+ Alertas de Datos
"""
import pandas as pd
import numpy as np
import warnings
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

df = pd.read_pickle('df_clean.pkl')
prod_agg = pd.read_pickle('prod_agg.pkl')
cli_real = pd.read_pickle('cli_real.pkl')

MAX_DATE = df['Fecha'].max()
MIN_DATE = df['Fecha'].min()

FONT_NAME = 'Arial'
COLOR_HEADER = '1F4E78'
COLOR_TOTAL = 'D9E1F2'
COLOR_KPI = '203864'
COLOR_GROW = 'C6EFCE'
COLOR_DROP = 'FFC7CE'

thin = Side(border_style='thin', color='BFBFBF')
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

def fmt_header(cell):
    cell.font = Font(name=FONT_NAME, size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', start_color=COLOR_HEADER)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border_all

def fmt_title(cell):
    cell.font = Font(name=FONT_NAME, size=14, bold=True, color=COLOR_HEADER)
    cell.alignment = Alignment(horizontal='left', vertical='center')

def fmt_kpi_label(cell):
    cell.font = Font(name=FONT_NAME, size=10, bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', start_color=COLOR_KPI)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_all

def fmt_kpi_value(cell):
    cell.font = Font(name=FONT_NAME, size=14, bold=True, color=COLOR_HEADER)
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border_all

def fmt_cell(cell, num_fmt=None, bold=False):
    cell.font = Font(name=FONT_NAME, size=10, bold=bold)
    cell.border = border_all
    if num_fmt:
        cell.number_format = num_fmt

def fmt_total_row(cell):
    cell.font = Font(name=FONT_NAME, size=10, bold=True)
    cell.fill = PatternFill('solid', start_color=COLOR_TOTAL)
    cell.border = border_all

def autosize(ws, widths):
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

CURRENCY = '$#,##0;($#,##0);-'
INT_FMT = '#,##0;(#,##0);-'
PCT_FMT = '0.0%;(0.0%);-'

wb = load_workbook('Analisis_Global_Family_Mayo2026.xlsx')

# =========================================================
# HOJA 10: DASHBOARD DATOS
# =========================================================
print('Construyendo HOJA 10: Dashboard Datos...')
ws = wb.create_sheet('10. Dashboard Datos')

ws['A1'] = 'DASHBOARD — KPIs CONSOLIDADOS'
fmt_title(ws['A1'])
ws.merge_cells('A1:F1')

ws['A2'] = f'Datos al {MAX_DATE.strftime("%Y-%m-%d")} | Período: {MIN_DATE.strftime("%Y-%m-%d")} a {MAX_DATE.strftime("%Y-%m-%d")}'
ws['A2'].font = Font(name=FONT_NAME, size=10, italic=True, color='595959')

# Cálculos
total_global = df['Total'].sum()
unidades_tot = df['Cantidad'].sum()
n_clientes_real = cli_real['Cliente'].nunique()
n_skus_activos = df[df['Fecha'] >= MAX_DATE - pd.Timedelta(days=60)]['Cod Interno'].nunique()

# Ventanas
def ventas_ventana(dias):
    fecha_lim = MAX_DATE - pd.Timedelta(days=dias)
    return df[df['Fecha'] > fecha_lim]['Total'].sum()

def unidades_ventana(dias):
    fecha_lim = MAX_DATE - pd.Timedelta(days=dias)
    return df[df['Fecha'] > fecha_lim]['Cantidad'].sum()

def facturas_ventana(dias):
    fecha_lim = MAX_DATE - pd.Timedelta(days=dias)
    return df[df['Fecha'] > fecha_lim]['Factura'].nunique()

v30 = ventas_ventana(30); v60 = ventas_ventana(60); v90 = ventas_ventana(90)
u30 = unidades_ventana(30); u60 = unidades_ventana(60); u90 = unidades_ventana(90)
f30 = facturas_ventana(30); f60 = facturas_ventana(60); f90 = facturas_ventana(90)

# --- KPIs visuales grid 2x4 ---
row = 4
kpis_visual = [
    ('FACTURACIÓN TOTAL', f'${total_global:,.0f}'),
    ('UNIDADES TOTALES', f'{unidades_tot:,.0f}'),
    ('CLIENTES ACTIVOS REALES', f'{n_clientes_real:,.0f}'),
    ('SKUs ACTIVOS (últ. 60d)', f'{n_skus_activos:,.0f}'),
]
for i, (label, val) in enumerate(kpis_visual):
    col = 1 + (i % 4) * 2
    # label
    c1 = ws.cell(row=row, column=col, value=label); fmt_kpi_label(c1)
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+1)
    # value
    c2 = ws.cell(row=row+1, column=col, value=val); fmt_kpi_value(c2)
    ws.merge_cells(start_row=row+1, start_column=col, end_row=row+2, end_column=col+1)

for r in range(row, row+3):
    ws.row_dimensions[r].height = 22

row += 5
# --- Ventas por ventana ---
ws.cell(row=row, column=1, value='VENTAS POR VENTANA TEMPORAL').font = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_HEADER)
row += 1
headers = ['Ventana', 'Facturación', 'Unidades', 'Nº Facturas', 'Ticket Promedio']
for i, h in enumerate(headers):
    c = ws.cell(row=row, column=i+1, value=h); fmt_header(c)
row += 1
for label, v, u, f in [('Últimos 30 días', v30, u30, f30), ('Últimos 60 días', v60, u60, f60), ('Últimos 90 días', v90, u90, f90)]:
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=int(v))
    ws.cell(row=row, column=3, value=int(u))
    ws.cell(row=row, column=4, value=int(f))
    ws.cell(row=row, column=5, value=f'=IFERROR(B{row}/D{row},0)')
    for col in range(1, 6):
        fmt_cell(ws.cell(row=row, column=col))
    ws.cell(row=row, column=2).number_format = CURRENCY
    ws.cell(row=row, column=3).number_format = INT_FMT
    ws.cell(row=row, column=4).number_format = INT_FMT
    ws.cell(row=row, column=5).number_format = CURRENCY
    row += 1

# Nota YoY
row += 1
ws.cell(row=row, column=1, value='Nota: no se incluye comparativo año vs año porque el histórico cubre solo ~7 meses (Oct 2025 – May 2026).').font = Font(name=FONT_NAME, size=9, italic=True, color='C00000')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)

# --- Top 5 productos ---
row += 3
ws.cell(row=row, column=1, value='TOP 5 PRODUCTOS POR FACTURACIÓN').font = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_HEADER)
row += 1
top5_p = prod_agg.nlargest(5, 'Ingresos')
headers = ['#', 'Producto', 'Categoría', 'Ingresos', '% Total']
for i, h in enumerate(headers):
    c = ws.cell(row=row, column=i+1, value=h); fmt_header(c)
row += 1
for i, (_, r) in enumerate(top5_p.iterrows()):
    ws.cell(row=row, column=1, value=i+1)
    ws.cell(row=row, column=2, value=r['Descripcion Producto'])
    ws.cell(row=row, column=3, value=r['CategoriaProducto'])
    ws.cell(row=row, column=4, value=int(r['Ingresos']))
    ws.cell(row=row, column=5, value=f'=D{row}/{int(total_global)}')
    for col in range(1, 6):
        fmt_cell(ws.cell(row=row, column=col))
    ws.cell(row=row, column=4).number_format = CURRENCY
    ws.cell(row=row, column=5).number_format = PCT_FMT
    row += 1

# --- Top 5 clientes ---
row += 2
ws.cell(row=row, column=1, value='TOP 5 CLIENTES (excl. CONSUMIDOR FINAL)').font = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_HEADER)
row += 1
top5_c = cli_real.nlargest(5, 'Ingresos')
headers = ['#', 'Cliente', 'Facturación', 'Nº Facturas', '% Total real']
for i, h in enumerate(headers):
    c = ws.cell(row=row, column=i+1, value=h); fmt_header(c)
row += 1
total_real = cli_real['Ingresos'].sum()
for i, (_, r) in enumerate(top5_c.iterrows()):
    ws.cell(row=row, column=1, value=i+1)
    ws.cell(row=row, column=2, value=r['Cliente'])
    ws.cell(row=row, column=3, value=int(r['Ingresos']))
    ws.cell(row=row, column=4, value=int(r['Facturas']))
    ws.cell(row=row, column=5, value=f'=C{row}/{int(total_real)}')
    for col in range(1, 6):
        fmt_cell(ws.cell(row=row, column=col))
    ws.cell(row=row, column=3).number_format = CURRENCY
    ws.cell(row=row, column=4).number_format = INT_FMT
    ws.cell(row=row, column=5).number_format = PCT_FMT
    row += 1

# --- Top 3 categorías ---
row += 2
ws.cell(row=row, column=1, value='TOP 3 CATEGORÍAS POR FACTURACIÓN').font = Font(name=FONT_NAME, size=12, bold=True, color=COLOR_HEADER)
row += 1
top3_cat = df.groupby('CategoriaProducto')['Total'].sum().nlargest(3)
headers = ['#', 'Categoría', 'Facturación', '% Total']
for i, h in enumerate(headers):
    c = ws.cell(row=row, column=i+1, value=h); fmt_header(c)
row += 1
for i, (cat, v) in enumerate(top3_cat.items()):
    ws.cell(row=row, column=1, value=i+1)
    ws.cell(row=row, column=2, value=cat)
    ws.cell(row=row, column=3, value=int(v))
    ws.cell(row=row, column=4, value=f'=C{row}/{int(total_global)}')
    for col in range(1, 5):
        fmt_cell(ws.cell(row=row, column=col))
    ws.cell(row=row, column=3).number_format = CURRENCY
    ws.cell(row=row, column=4).number_format = PCT_FMT
    row += 1

autosize(ws, {'A': 5, 'B': 55, 'C': 28, 'D': 18, 'E': 14, 'F': 14})

# =========================================================
# HOJA: NOTAS METODOLÓGICAS
# =========================================================
print('Construyendo hoja Notas Metodológicas...')
ws = wb.create_sheet('Notas Metodologicas')

ws['A1'] = 'NOTAS METODOLÓGICAS'
fmt_title(ws['A1'])

notas = [
    ('PERÍODO ANALIZADO', f'Datos comprendidos entre {MIN_DATE.strftime("%Y-%m-%d")} y {MAX_DATE.strftime("%Y-%m-%d")} (~7 meses).'),
    ('FUENTE DE DATOS', 'Archivo "Venta Global Family al 21 mayo 2026.xlsx", hoja "Venta" — 38,961 líneas de factura.'),
    ('LIMPIEZA APLICADA',
        '• Caracteres mal codificados (carácter "�") sustituidos por la vocal acentuada correcta donde se pudo identificar.\n'
        '• Nombres de Cliente y Descripción de Producto normalizados a MAYÚSCULAS y con espacios colapsados.\n'
        '• Se filtraron registros con Total ≤ 0 o Cantidad ≤ 0 (0 registros excluidos en este archivo).\n'
        '• El conteo de facturas únicas se hace por el número de Factura — cada factura suele tener varias líneas.'),
    ('MÉTRICA DE FACTURACIÓN',
        'Se usa la columna "Total" (con impuestos) como medida de ingresos en todos los cálculos. No se usa "Base" ni "Total Venta".'),
    ('CATEGORÍAS DE PRODUCTO',
        'La columna "Categoria" del archivo original contiene PROVEEDOR/MARCA (NESTLE, JOHNSONS, SOFTYS, etc.) y NO una categoría de producto. \n'
        'Por eso se genera la HOJA 2 "Categorías Producto" con categorías inferidas a partir de palabras clave en la descripción del producto. \n'
        'La HOJA 2b muestra el agrupamiento por proveedor/marca tal cual viene en el archivo.\n\n'
        'Reglas de categorización (primer match gana, en este orden):\n'
        '  PAÑALES: PAÑAL/PANAL, PAMPERS, HUGGIES, WINNY, BABYSEC, PEQUEÑIN, PUFFIES\n'
        '  TOALLITAS HÚMEDAS Y PAÑITOS: TOALLITA, TOALLA HUMED, PAÑITO/PANITO, WIPE\n'
        '  ESTUCHES Y KITS REGALO: ESTUCHE, ARRURRU, BABY SHOWER, KIT REGALO\n'
        '  BAÑO BEBÉ: BAÑERA, BACIN, ORINAL, BACINILLA\n'
        '  MOBILIARIO INFANTIL: CORRAL, CAMPING, SILLA COMEDOR, CUNA, COLCHÓN, MOISÉS, ESCRITORIO, MESA INF, TOCADOR\n'
        '  CAMINADORES Y MONTABLES: CARRO MONTABLE, CAMINADOR, ANDADOR, TRICICLO\n'
        '  COCHES Y PORTEO: COCHE BEBÉ, CARRIOLA, SILLA AUTO, CANGURO, CARGADOR (porteo)\n'
        '  ALIMENTACIÓN INFANTIL: FORMULA, NAN, NIDO, NESTOGENO, ENFAMIL, COMPOTA, PAPILLA, CEREAL INF\n'
        '  BIBERONES Y EXTRACTORES: BIBERON, TETERO, CHUPO, CHUPETE, EXTRACTOR LECHE\n'
        '  CUIDADO DENTAL / FEMENINO / HIGIENE PERSONAL / CREMAS / etc. (ver código fuente)\n\n'
        'Productos sin match → categoría OTROS. En este archivo: 578 SKUs / 9.2% de la facturación.'),
    ('CLIENTES GENÉRICOS / CONSUMIDOR FINAL',
        'Se identifica como "ClienteGenerico" a los registros cuyo nombre contiene CONSUMIDOR FINAL, CLIENTE OCASIONAL, '
        'VENTA MOSTRADOR, etc. En este archivo solo se detectó "CONSUMIDOR FINAL". Estos clientes se EXCLUYEN del Top 20, '
        'de la segmentación ABC y de la concentración de clientes, pero SÍ se incluyen en los totales generales del negocio.'),
    ('SEGMENTACIÓN ABC DE CLIENTES',
        'Los clientes reales (excl. CONSUMIDOR FINAL) se ordenan por facturación descendente y se asigna:\n'
        '  A = Top 20% (clientes más valiosos)\n'
        '  B = Siguiente 30%\n'
        '  C = Restante 50%\n'
        'La definición se basa en CANTIDAD de clientes, no en valor acumulado — el % de facturación acumulado se reporta por separado.'),
    ('ALERTA DE RECUPERACIÓN',
        'Clientes "inactivos" = clientes reales con ≥3 facturas en su historia (compradores recurrentes) y sin compras en los últimos 60 días o más. '
        'Pintados en naranja si llevan 60-90 días sin comprar, rojo si llevan más de 90.'),
    ('PRODUCTOS ESTRATÉGICOS',
        '• Estrella: aparece en ambos rankings Top 20 (facturación + rotación)\n'
        '• Vaca lechera: en Top 20 facturación pero NO en Top 20 rotación (margen alto, volumen bajo)\n'
        '• Volumen: en Top 20 rotación pero NO en Top 20 facturación (alto movimiento, precio bajo)\n'
        '• Declive: top 10 con mayor caída % comparando últimos 3 meses vs 3 meses anteriores. Solo se consideran productos con ≥$500.000 de venta total en esos 6 meses.\n'
        '• Ascenso: top 10 con mayor crecimiento %, mismas condiciones.\n'
        '• Cola larga: productos con frecuencia de aparición en facturas < 1 vez/mes en promedio. Candidatos a descontinuar.'),
    ('PARETO 80/20',
        '"Productos que generan el 80% de la facturación" = los N primeros productos ordenados por ingresos cuya suma acumulada llega al 80% del total.'),
    ('CROSS-SELLING',
        'Calculado sobre los top 200 productos por facturación. Se cuentan los pares (A, B) que aparecen en la misma factura. '
        'Se limita a facturas con ≤30 productos distintos para evitar combinaciones explosivas. Se reportan los 15 pares más frecuentes.'),
    ('OPORTUNIDAD DE RECOMENDACIÓN A→B',
        'Productos comprados por ≥30% de los clientes del segmento A pero por <10% de los clientes del segmento B. '
        'Indica un producto que los clientes premium consideran imprescindible y que los clientes medianos podrían adoptar.'),
    ('TENDENCIA',
        'Calculada por regresión lineal simple sobre la facturación de los últimos 6 meses. Clasificación: '
        'CRECIENTE si pendiente positiva y cambio inicial→final > +5%; DECRECIENTE si < -5%; ESTABLE en otro caso.'),
    ('DATOS NO INCLUIDOS / NO CALCULADOS',
        '• Comparativo año vs año (YoY): el histórico cubre solo ~7 meses, no hay año previo.\n'
        '• Análisis de margen / rentabilidad por producto: el archivo no contiene costo unitario.\n'
        '• Análisis de devoluciones: no hay columna de tipo "devolución" diferenciada.\n'
        '• Análisis por vendedor / canal: no existe columna de vendedor o canal en el archivo.'),
]
row = 3
for titulo, texto in notas:
    ws.cell(row=row, column=1, value=titulo).font = Font(name=FONT_NAME, size=11, bold=True, color='FFFFFF')
    ws.cell(row=row, column=1).fill = PatternFill('solid', start_color=COLOR_HEADER)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    row += 1
    c = ws.cell(row=row, column=1, value=texto)
    c.font = Font(name=FONT_NAME, size=10)
    c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    # estimar alto de fila por número de líneas
    n_lines = max(1, texto.count('\n') + 1 + len(texto) // 90)
    ws.row_dimensions[row].height = max(20, min(n_lines * 15, 250))
    row += 2

autosize(ws, {'A': 50, 'B': 50, 'C': 50})

# =========================================================
# HOJA: ALERTAS DE DATOS
# =========================================================
print('Construyendo hoja Alertas de Datos...')
ws = wb.create_sheet('Alertas de Datos')

ws['A1'] = 'ALERTAS DE DATOS — incidencias detectadas durante el procesamiento'
fmt_title(ws['A1'])

alertas = []

# 1) Caracteres mal codificados
total_lineas_orig = 38961
n_bad_chars = df['Cliente'].astype(str).str.contains('�').sum() + df['Descripcion Producto'].astype(str).str.contains('�').sum()
alertas.append((
    'BAJA' if n_bad_chars == 0 else 'MEDIA',
    'Codificación de caracteres',
    f'Algunos campos del archivo original contienen el carácter "�" (encoding ANSI/UTF-8 mezclado). Se reemplazó por "ó" o "é" según contexto en columnas Tipo Fact., Documento. Tras limpieza quedaron {n_bad_chars} ocurrencias en Cliente/Producto.'
))

# 2) Nulos en cod barras
n_null_barras = df['Cod. Barras'].isna().sum()
pct_null_barras = n_null_barras / len(df) * 100
alertas.append((
    'BAJA',
    'Códigos de barras faltantes',
    f'{n_null_barras:,} registros ({pct_null_barras:.1f}%) no tienen código de barras. No afecta análisis (usamos Cod Interno como llave de producto).'
))

# 3) Nulos en forma de pago
n_null_pago = df['Forma de Pago'].isna().sum()
pct_null_pago = n_null_pago / len(df) * 100
alertas.append((
    'BAJA',
    'Forma de Pago faltante',
    f'{n_null_pago:,} registros ({pct_null_pago:.1f}%) no tienen forma de pago registrada. No se usa esta columna en el análisis principal.'
))

# 4) Productos con valor unitario muy variable
prod_price_var = df.groupby('Cod Interno').agg(
    pmin=('Valor Unidad', 'min'),
    pmax=('Valor Unidad', 'max'),
    pmean=('Valor Unidad', 'mean'),
    n=('Valor Unidad', 'count'),
    desc=('Descripcion Producto', 'first')
)
prod_price_var['var_ratio'] = prod_price_var['pmax'] / prod_price_var['pmin'].replace(0, np.nan)
prod_var_alta = prod_price_var[(prod_price_var['var_ratio'] > 3) & (prod_price_var['n'] >= 5)]
alertas.append((
    'MEDIA' if len(prod_var_alta) > 50 else 'BAJA',
    'Productos con precio unitario altamente variable',
    f'{len(prod_var_alta)} productos tienen precios cuyo máximo es ≥3x el mínimo (con al menos 5 ventas). Puede indicar promociones, errores de captura o cambios de empaque.'
))

# 5) Facturas grandes anómalas
fact_total = df.groupby('Factura')['Total'].sum()
mediana_fact = fact_total.median()
fact_grandes = fact_total[fact_total > mediana_fact * 50]
alertas.append((
    'BAJA' if len(fact_grandes) < 10 else 'MEDIA',
    'Facturas con monto extremadamente alto',
    f'{len(fact_grandes)} facturas tienen un monto superior a 50x la mediana (mediana = ${mediana_fact:,.0f}). Pueden ser legítimas (clientes grandes) pero conviene revisarlas.'
))

# 6) Productos asignados a OTROS
n_otros = (df['CategoriaProducto'] == 'OTROS').sum()
pct_otros = n_otros / len(df) * 100
ing_otros = df[df['CategoriaProducto'] == 'OTROS']['Total'].sum()
pct_ing_otros = ing_otros / df['Total'].sum() * 100
alertas.append((
    'MEDIA',
    'Productos sin categoría asignada (OTROS)',
    f'{n_otros:,} líneas ({pct_otros:.1f}% de líneas) corresponden a productos que no pudieron clasificarse por keywords, equivalentes al {pct_ing_otros:.1f}% de la facturación (${ing_otros:,.0f}). Se requiere revisión manual para refinar las reglas.'
))

# 7) Verificar duplicados de líneas idénticas (mismo producto/cantidad/factura repetido)
dup = df.duplicated(subset=['Factura', 'Cod Interno', 'Cantidad', 'Valor Unidad']).sum()
alertas.append((
    'BAJA' if dup < 100 else 'MEDIA',
    'Posibles líneas duplicadas',
    f'{dup:,} filas son duplicados exactos de (Factura, Cod Interno, Cantidad, Valor Unidad). Puede ser legítimo (mismo producto cargado dos veces en la misma factura) pero conviene auditar.'
))

# Render
headers = ['Severidad', 'Tipo de alerta', 'Descripción']
row = 3
for i, h in enumerate(headers):
    c = ws.cell(row=row, column=i+1, value=h); fmt_header(c)
row += 1
sev_colors = {'ALTA': 'FFC7CE', 'MEDIA': 'FFEB9C', 'BAJA': 'C6EFCE'}
for sev, tipo, desc in alertas:
    c1 = ws.cell(row=row, column=1, value=sev)
    c1.font = Font(name=FONT_NAME, size=10, bold=True)
    c1.fill = PatternFill('solid', start_color=sev_colors.get(sev, 'FFFFFF'))
    c1.alignment = Alignment(horizontal='center', vertical='center')
    c1.border = border_all
    c2 = ws.cell(row=row, column=2, value=tipo); fmt_cell(c2, bold=True)
    c3 = ws.cell(row=row, column=3, value=desc); fmt_cell(c3)
    c3.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    n_lines = max(1, len(desc) // 80 + 1)
    ws.row_dimensions[row].height = max(40, n_lines * 18)
    row += 1

autosize(ws, {'A': 14, 'B': 35, 'C': 80})

# =========================================================
# Reordenar hojas para presentación lógica
# =========================================================
ws_order = [
    '1. Resumen Ejecutivo',
    '2. Categorias Producto',
    '2b. Categorias Proveedor',
    '3. Top20 por Facturacion',
    '4. Top20 por Rotacion',
    '5. Analisis Clientes',
    '6. Temporal Estacionalidad',
    '7. Productos Estrategicos',
    '8. Concentracion Riesgo',
    '9. Oportunidades',
    '10. Dashboard Datos',
    'Notas Metodologicas',
    'Alertas de Datos',
]
# Reset order
existing = wb.sheetnames
new_order = [s for s in ws_order if s in existing]
# Apply
wb._sheets = [wb[name] for name in new_order]

# Set tab colors
tab_colors = {
    '1. Resumen Ejecutivo': '1F4E78',
    '2. Categorias Producto': '2E75B6',
    '2b. Categorias Proveedor': '2E75B6',
    '3. Top20 por Facturacion': '2E75B6',
    '4. Top20 por Rotacion': '2E75B6',
    '5. Analisis Clientes': '2E75B6',
    '6. Temporal Estacionalidad': '2E75B6',
    '7. Productos Estrategicos': '70AD47',
    '8. Concentracion Riesgo': 'C00000',
    '9. Oportunidades': '70AD47',
    '10. Dashboard Datos': '1F4E78',
    'Notas Metodologicas': '595959',
    'Alertas de Datos': 'BF8F00',
}
for name, color in tab_colors.items():
    if name in wb.sheetnames:
        wb[name].sheet_properties.tabColor = color

wb.save('Analisis_Global_Family_Mayo2026.xlsx')
print('Hoja 10 + Notas + Alertas listas. Archivo guardado.')
print(f'Hojas finales: {wb.sheetnames}')
