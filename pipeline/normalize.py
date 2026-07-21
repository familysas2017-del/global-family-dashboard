"""
Normaliza cada fuente cruda en tablas planas listas para consumo analítico.
Genera CSVs en data/clean/. Tolerante a fallas por fuente.
"""
from __future__ import annotations
import logging
import re
import unicodedata
from pathlib import Path

import numpy as np
import pandas as pd

from config import (
    RAW_DIR, CLEAN_DIR, EXPECTED_FILES, CATEGORIZATION_RULES,
    GASTOS_FIJOS, GASTOS_VARIABLES, GASTOS_NO_OPERACIONALES, GASTOS_COSTO_IMPORTACION,
)

log = logging.getLogger(__name__)


# =========================================================
# UTILIDADES
# =========================================================
_RAW_ANCHORS = {
    "ventas_hist":       ["ventas", "junio2024"],
    "analisis_num":      ["family", "analisis"],
    "gastos":            ["gastos", "global"],
    "cartera":           ["cartera"],
    "cxp_nacional":      ["cuentas", "pagar", "nacional"],
    "cxp_internacional": ["cuentas", "pagar", "int"],
    "pagos_jeison":      ["pagos", "jeison"],
}


def _raw_path(alias: str) -> Path:
    """Busca el archivo real en data/raw/ por match ESTRICTO con anclas."""
    import unicodedata
    def _norm(s: str) -> str:
        return unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii").lower()

    expected = EXPECTED_FILES[alias]
    p = RAW_DIR / expected
    if p.exists():
        return p
    anchors = [_norm(a) for a in _RAW_ANCHORS.get(alias, [_norm(expected).split()[0]])]
    candidates = []
    for f in RAW_DIR.iterdir():
        name_lc = _norm(f.name)
        if all(a in name_lc for a in anchors):
            candidates.append(f)
    # excluir cross-match nacional/internacional
    if alias == "cxp_internacional":
        candidates = [c for c in candidates if "nacional" not in _norm(c.name)]
    if alias == "cxp_nacional":
        candidates = [c for c in candidates if "int.xlsx" not in _norm(c.name)
                      and "internacional" not in _norm(c.name)]
    if candidates:
        return candidates[0]
    raise FileNotFoundError(f"No hay archivo raw para alias={alias} (esperado {expected})")


def snake_case(s: str) -> str:
    """Convierte a snake_case sin tildes."""
    s = str(s)
    # quitar tildes
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = s.replace("$", "").replace("[", "").replace("]", "")
    s = re.sub(r"[^\w\s]", "_", s)
    s = re.sub(r"\s+", "_", s.strip())
    s = re.sub(r"_+", "_", s)
    return s.strip("_").lower()


def normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [snake_case(c) for c in df.columns]
    return df


def clean_encoding(s):
    if pd.isna(s) or not isinstance(s, str):
        return s
    # heurística: el � apareció donde iba una vocal acentuada — usamos "ó" como reemplazo genérico
    # (cadenas específicas ya sabidas se manejan aparte)
    return s.replace("Cr�dito", "Crédito").replace("Remisi�n", "Remisión") \
            .replace("Electr�nica", "Electrónica").replace("�", "o")


def to_numeric_safe(s):
    """Convierte a numérico, #N/A/#DIV0! → NaN."""
    return pd.to_numeric(s, errors="coerce")


def categorize_product(desc: str) -> str:
    if not isinstance(desc, str) or not desc:
        return "OTROS"
    d = desc.upper()
    for cat, keys in CATEGORIZATION_RULES:
        for k in keys:
            if k in d:
                return cat
    return "OTROS"


def _save(df: pd.DataFrame, name: str) -> Path:
    p = CLEAN_DIR / name
    df.to_csv(p, index=False, encoding="utf-8")
    log.info("  → %s (%d filas × %d cols)", name, len(df), len(df.columns))
    return p


# =========================================================
# A) VENTAS TRANSACCIONALES (histórico grande, hoja Reporte_detalle_facturas_por_pr)
# =========================================================
def normalize_ventas() -> pd.DataFrame:
    log.info("[A] Ventas transaccionales (~119k líneas)…")
    p = _raw_path("ventas_hist")
    df = pd.read_excel(p, sheet_name="Reporte_detalle_facturas_por_pr")
    df = normalize_cols(df)
    # Fechas: día en `fecha_factura`, más `mes`, `ano`
    df["ano"] = to_numeric_safe(df["ano"])
    df["mes"] = to_numeric_safe(df["mes"])
    df["dia"] = to_numeric_safe(df["fecha_factura"])
    # Construir fecha componente por componente (tolerante a NaN)
    df["fecha"] = pd.to_datetime(
        dict(year=df["ano"], month=df["mes"], day=df["dia"]),
        errors="coerce",
    )
    # texto
    for c in ("cliente", "descripcion_producto", "categoria", "tipo_fact", "tipo_de_documento", "forma_de_pago"):
        if c in df.columns:
            df[c] = df[c].map(clean_encoding).astype("string").str.strip()

    # normalización de nombres (uppercase, espacios simples) para cliente y producto
    for c in ("cliente", "descripcion_producto"):
        df[c] = df[c].str.upper().str.replace(r"\s+", " ", regex=True)

    # numéricos
    for c in ("cantidad", "valor_unidad", "total_venta", "base", "impuesto",
              "ico", "icui", "ibua", "descuento", "total", "costo_promedio", "p_compra"):
        if c in df.columns:
            df[c] = to_numeric_safe(df[c])

    # rename final para consistencia
    df = df.rename(columns={
        "categoria": "proveedor_marca",
        "tipo_fact": "tipo_factura",
        "tipo_de_documento": "tipo_documento",
        "ano": "anio",
        "p_compra": "precio_compra",
    })

    # calculadas
    df["costo_total_linea"] = (df["cantidad"] * df["costo_promedio"]).round(2)
    df["margen_bruto_linea"] = (df["total_venta"] - df["costo_total_linea"]).round(2)
    df["margen_bruto_pct"] = np.where(
        (df["total_venta"].fillna(0) > 0),
        (df["margen_bruto_linea"] / df["total_venta"]).round(4),
        np.nan,
    )
    df["anio_mes"] = df["fecha"].dt.to_period("M").astype("string")
    df["dia_semana"] = df["fecha"].dt.day_name()

    # categoría producto (18 categorías bebé)
    df["categoria_producto"] = df["descripcion_producto"].map(categorize_product)

    # filtro sanidad: descarta filas sin fecha o con total<=0/cantidad<=0
    n_pre = len(df)
    df = df[df["fecha"].notna() & (df["total"].fillna(0) > 0) & (df["cantidad"].fillna(0) > 0)]
    log.info("  Filtradas %d filas (de %d → %d) por fecha nula / total<=0 / cantidad<=0",
             n_pre - len(df), n_pre, len(df))

    # cliente genérico
    df["cliente_generico"] = df["cliente"].fillna("").str.contains(
        r"CONSUMIDOR FINAL|CLIENTE OCASIONAL|CLIENTE GENERAL|PUBLICO GENERAL|VENTA MOSTRADOR",
        regex=True, na=False
    )

    _save(df, "fact_ventas.csv")
    return df


# =========================================================
# B) INVENTARIO (histórico grande, hoja Inv )
# =========================================================
def normalize_inventario() -> pd.DataFrame:
    log.info("[B] Inventario (960 SKUs)…")
    p = _raw_path("ventas_hist")
    df = pd.read_excel(p, sheet_name="Inv ")
    df = normalize_cols(df)
    df["producto"] = df["producto"].map(clean_encoding).str.upper().str.strip()
    df["sistema"] = to_numeric_safe(df["sistema"])
    df["costo"]   = to_numeric_safe(df["costo"])
    df = df.rename(columns={"sistema": "stock_actual", "costo": "costo_total_inventario"})
    df["costo_unitario_inventario"] = np.where(
        df["stock_actual"].fillna(0) > 0,
        (df["costo_total_inventario"] / df["stock_actual"]).round(2),
        np.nan,
    )
    df["categoria_producto"] = df["producto"].map(categorize_product)
    _save(df, "dim_inventario.csv")
    return df


# =========================================================
# C) CARTERA
# =========================================================
def _parse_fecha_mixta(v):
    """Formato mixto: '2025-11-18 00:00:00' o '17/Ene/2026'."""
    if pd.isna(v):
        return pd.NaT
    if isinstance(v, (pd.Timestamp,)):
        return v
    try:
        return pd.to_datetime(v)
    except Exception:
        pass
    # 17/Ene/2026 style
    meses_es = {"ene":1,"feb":2,"mar":3,"abr":4,"may":5,"jun":6,"jul":7,"ago":8,"sep":9,"oct":10,"nov":11,"dic":12}
    m = re.match(r"^\s*(\d{1,2})/([A-Za-z]{3})/(\d{4})\s*$", str(v))
    if m:
        d, mes_txt, y = m.groups()
        mm = meses_es.get(mes_txt.lower()[:3])
        if mm:
            try:
                return pd.Timestamp(year=int(y), month=mm, day=int(d))
            except Exception:
                return pd.NaT
    return pd.NaT


def normalize_cartera(fecha_hoy: pd.Timestamp) -> pd.DataFrame:
    log.info("[C] Cartera…")
    p = _raw_path("cartera")
    # headers en fila 2 → header=1
    raw = pd.read_excel(p, sheet_name=0, header=1)
    # las últimas columnas son un mini-resumen aging que no nos interesa
    keep = [c for c in ["URL", "FACTURA", "CLIENTE", "FECHA", "FECHA LIMITE DE PAGO", "TOTAL FACTURA", "RESTA", "DIAS CARTERA"] if c in raw.columns]
    df = raw[keep].copy()
    df = normalize_cols(df)
    df = df.rename(columns={
        "url": "tipo_documento",
        "factura": "num_factura",
        "cliente": "cliente",
        "fecha": "fecha_factura",
        "fecha_limite_de_pago": "fecha_limite_pago",
        "total_factura": "total_factura",
        "resta": "saldo_pendiente",
        "dias_cartera": "dias_cartera_reportado",
    })
    df = df[df["num_factura"].notna()]
    df["cliente"] = df["cliente"].map(clean_encoding).astype("string").str.strip()
    df["fecha_factura"] = df["fecha_factura"].map(_parse_fecha_mixta)
    df["fecha_limite_pago"] = df["fecha_limite_pago"].map(_parse_fecha_mixta)
    df["total_factura"] = to_numeric_safe(df["total_factura"])
    df["saldo_pendiente"] = to_numeric_safe(df["saldo_pendiente"])
    df["dias_vencido"] = (fecha_hoy - df["fecha_limite_pago"]).dt.days
    df["dias_vencido"] = df["dias_vencido"].fillna(0).astype(int)

    def tramo(dv):
        if dv <= 0: return "corriente"
        if dv <= 30: return "1-30"
        if dv <= 60: return "31-60"
        if dv <= 90: return "61-90"
        return ">90"
    df["tramo_aging"] = df["dias_vencido"].map(tramo)
    _save(df, "fact_cartera.csv")
    return df


# =========================================================
# D) CxP NACIONAL (Compras + Abonos)
# =========================================================
def normalize_cxp_nacional(fecha_hoy: pd.Timestamp) -> pd.DataFrame:
    log.info("[D] CxP Nacional…")
    p = _raw_path("cxp_nacional")
    compras = pd.read_excel(p, sheet_name="Compras")
    abonos = pd.read_excel(p, sheet_name="Abonos")
    compras = normalize_cols(compras)
    abonos  = normalize_cols(abonos)
    compras = compras[compras["proveedor"].notna()].copy()
    abonos  = abonos[abonos["proveedor"].notna()].copy()

    # numéricos
    for c in ("valor_base", "valor_iva", "valor_total", "plazo", "total_abonos", "saldo_a_pagar"):
        if c in compras.columns:
            compras[c] = to_numeric_safe(compras[c])
    compras["fecha_factura"] = pd.to_datetime(compras["fecha_factura"], errors="coerce")
    compras["fecha_vencimiento"] = pd.to_datetime(compras["fecha_vencimiento"], errors="coerce")

    # Llave: en compras la factura real está en "documento" (no "documento_1" — esa siempre es 0).
    # En abonos, la factura viene en "factura". Ambas son tipo mixto (int/str) → normalizamos a str.
    keycol_compras = "documento"
    keycol_abonos  = "factura" if "factura" in abonos.columns else "num_factura"

    abonos["abonos"] = to_numeric_safe(abonos["abonos"])
    # Normalizar llaves a string (los int 397 y str '397' deben cruzar)
    compras["_llave"]  = compras[keycol_compras].astype(str).str.strip().str.upper()
    abonos["_llave"]   = abonos[keycol_abonos].astype(str).str.strip().str.upper()
    # También normalizar proveedor (case + trim)
    compras["_prov_key"] = compras["proveedor"].astype(str).str.strip().str.upper()
    abonos["_prov_key"]  = abonos["proveedor"].astype(str).str.strip().str.upper()

    abonos_g = (abonos.groupby(["_llave", "_prov_key"], dropna=False)
                       .agg(total_abonado=("abonos", "sum"),
                            n_abonos=("abonos", "count"),
                            ultimo_abono=("fecha_abono", "max"))
                       .reset_index())

    df = compras.merge(abonos_g, on=["_llave", "_prov_key"], how="left", suffixes=("", "_ab"))
    df["total_abonado"] = df["total_abonado"].fillna(0)
    df = df.drop(columns=["_llave", "_prov_key"])
    df["saldo_pendiente"] = (df["valor_total"] - df["total_abonado"]).round(2)
    df["dias_por_pagar"] = (df["fecha_vencimiento"] - fecha_hoy).dt.days
    df["dias_vencido"] = np.where(df["dias_por_pagar"] < 0, -df["dias_por_pagar"], 0)

    def tramo(row):
        if row["saldo_pendiente"] <= 0: return "pagada"
        d = row["dias_vencido"]
        if d == 0: return "al_dia"
        if d <= 30: return "vencida_1-30"
        if d <= 60: return "vencida_31-60"
        if d <= 90: return "vencida_61-90"
        return "vencida_>90"
    df["tramo_aging"] = df.apply(tramo, axis=1)

    _save(df, "fact_cxp_nacional.csv")
    return df


# =========================================================
# E) CxP INTERNACIONAL (Compras + Abonos + Giros)
# =========================================================
def normalize_cxp_internacional(fecha_hoy: pd.Timestamp) -> pd.DataFrame:
    log.info("[E] CxP Internacional…")
    p = _raw_path("cxp_internacional")
    compras = pd.read_excel(p, sheet_name="Compras")
    abonos  = pd.read_excel(p, sheet_name="Abonos")
    giros   = pd.read_excel(p, sheet_name="Giros")
    compras = normalize_cols(compras); compras = compras[compras["proveedor"].notna()].copy()
    abonos  = normalize_cols(abonos);  abonos  = abonos[abonos["proveedor"].notna()].copy()
    giros   = normalize_cols(giros)

    for c in ("valor_usd", "trm", "valor_en_sistema", "plazo",
              "total_abono_usd", "total_abono_pesos", "saldo_a_pagar_usd", "saldo_a_pagar", "trm_prom"):
        if c in compras.columns:
            compras[c] = to_numeric_safe(compras[c])
    compras["fecha_factura"]     = pd.to_datetime(compras["fecha_factura"], errors="coerce")
    compras["fecha_vencimiento"] = pd.to_datetime(compras["fecha_vencimiento"], errors="coerce")

    abonos["abonos"] = to_numeric_safe(abonos["abonos"])
    keycol_compras = "documento"
    keycol_abonos  = "factura" if "factura" in abonos.columns else "num_factura"
    compras["_llave"]    = compras[keycol_compras].astype(str).str.strip().str.upper()
    abonos["_llave"]     = abonos[keycol_abonos].astype(str).str.strip().str.upper()
    compras["_prov_key"] = compras["proveedor"].astype(str).str.strip().str.upper()
    abonos["_prov_key"]  = abonos["proveedor"].astype(str).str.strip().str.upper()

    abonos_g = abonos.groupby(["_llave", "_prov_key"], dropna=False).agg(
        abonos_recientes=("abonos", "sum"), n_abonos=("abonos", "count")
    ).reset_index()

    df = compras.merge(abonos_g, on=["_llave", "_prov_key"], how="left")
    df["abonos_recientes"] = df["abonos_recientes"].fillna(0)
    df = df.drop(columns=["_llave", "_prov_key"])
    # Total abonado priorizando el campo del propio archivo si viene; sino usamos abonos_g
    df["total_abonado_pesos"] = df.get("total_abono_pesos").fillna(df["abonos_recientes"])
    df["saldo_pendiente_cop"] = df.get("saldo_a_pagar").fillna(
        df["valor_en_sistema"] - df["total_abonado_pesos"]
    )
    df["saldo_pendiente_usd"] = df.get("saldo_a_pagar_usd")
    df["dias_por_pagar"] = (df["fecha_vencimiento"] - fecha_hoy).dt.days
    df["tipo_cxp"] = "internacional"

    _save(df, "fact_cxp_internacional.csv")
    return df


# =========================================================
# F) GASTOS (Base Datos)
# =========================================================
def _clasificar_gasto(tipo):
    if pd.isna(tipo):
        return ("desconocido", "operacional")
    t = str(tipo).strip()
    # normalización de acentos para lookup insensible
    def norm(s): return unicodedata.normalize("NFKD", s).encode("ascii","ignore").decode("ascii").lower()
    tn = norm(t)
    fx = "variable"
    op = "operacional"
    for f in GASTOS_FIJOS:
        if norm(f) in tn or tn in norm(f):
            fx = "fijo"; break
    if fx == "variable":
        for v in GASTOS_VARIABLES:
            if norm(v) in tn or tn in norm(v):
                fx = "variable"; break
    for n in GASTOS_NO_OPERACIONALES:
        if norm(n) in tn or tn in norm(n):
            op = "no_operacional"; break
    for c in GASTOS_COSTO_IMPORTACION:
        if norm(c) in tn or tn in norm(c):
            op = "costo_importacion"; break
    return (fx, op)


def normalize_gastos() -> pd.DataFrame:
    log.info("[F] Gastos…")
    p = _raw_path("gastos")
    df = pd.read_excel(p, sheet_name="Base Datos")
    df = normalize_cols(df)
    # cols esperadas: fecha, dia, mes, ano, tipo, concepto, valor, medio_de_pago, observaciones
    df = df.rename(columns={"ano": "anio", "valor": "valor_cop"})
    df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
    df["valor_cop"] = to_numeric_safe(df["valor_cop"])
    df = df[df["fecha"].notna() & (df["valor_cop"].fillna(0) > 0)].copy()

    clases = df["tipo"].map(_clasificar_gasto)
    df["clasificacion_fx"] = clases.map(lambda x: x[0])
    df["clasificacion_op"] = clases.map(lambda x: x[1])
    df["anio_mes"] = df["fecha"].dt.to_period("M").astype("string")
    _save(df, "fact_gastos.csv")
    return df


# =========================================================
# G) VENTAS POR VENDEDOR (Análisis Numérico, hojas mensuales)
# =========================================================
VENDEDORES = ["Jaime", "James", "Jhon", "Botero", "Eugenia"]
MES_A_NUM = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5, "JUNIO": 6,
    "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9, "OCTUBRE": 10,
    "NOVIEMBRE": 11, "DICIEMBRE": 12,
}


def normalize_ventas_vendedor() -> pd.DataFrame:
    log.info("[G] Ventas por vendedor…")
    p = _raw_path("analisis_num")
    from openpyxl import load_workbook
    wb = load_workbook(p, data_only=True)
    filas = []
    for sname in wb.sheetnames:
        s = sname.strip().upper()
        m = re.match(r"([A-ZÁÉÍÓÚÑ]+)\s+(\d{4})", s)
        if not m:
            continue
        mes_txt, anio = m.group(1), int(m.group(2))
        mes_num = MES_A_NUM.get(mes_txt)
        if not mes_num:
            continue
        ws = wb[sname]
        # Estructura conocida (validada con archivo real, openpyxl 1-indexed):
        # Fila 12: C2=Aseo   C3=TOTAL_ELEC C4=TOTAL_REM
        #          C6/7=Jaime  C9/10=James  C12/13=Jhon  C15/16=Botero  C18/19=Eugenia
        # Fila 19: Importados+Varios  (misma estructura de columnas)
        # Fila 24: TOTAL GLOBAL FAMILY (no lo usamos, se recalcula)
        col_map = {"Jaime": (6, 7), "James": (9, 10), "Jhon": (12, 13), "Botero": (15, 16), "Eugenia": (18, 19)}
        for cat_row, cat_name in [(12, "Aseo"), (19, "Importados y Varios")]:
            for v, (c_elec, c_rem) in col_map.items():
                v_elec = ws.cell(row=cat_row, column=c_elec).value
                v_rem  = ws.cell(row=cat_row, column=c_rem).value
                for tipo, valor in [("Electronica", v_elec), ("Remisiones", v_rem)]:
                    if valor and isinstance(valor, (int, float)) and valor > 0:
                        filas.append({
                            "vendedor": v, "mes": mes_num, "anio": anio,
                            "anio_mes": f"{anio}-{mes_num:02d}",
                            "categoria_venta": cat_name,
                            "tipo_documento": tipo,
                            "monto_cop": float(valor),
                        })
    df = pd.DataFrame(filas)
    _save(df, "fact_ventas_vendedor.csv")
    return df


# =========================================================
# H) PAGOS JEISON
# =========================================================
def normalize_pagos_jeison(fecha_hoy: pd.Timestamp) -> pd.DataFrame:
    log.info("[H] Pagos Jeison…")
    p = _raw_path("pagos_jeison")
    df = pd.read_excel(p, sheet_name=0)
    df = normalize_cols(df)
    # cols esperadas: cuota, fecha_cuota, descripcion, valor_capital, valor_cuota, abono, nuevo_saldo_a_capital, saldo
    for c in df.columns:
        if c in ("valor_capital", "valor_cuota", "abono", "nuevo_saldo_a_capital_saldo",
                 "nuevo_saldo_a_capital", "saldo", "cuota"):
            df[c] = to_numeric_safe(df[c])
    df["fecha_cuota"] = pd.to_datetime(df["fecha_cuota"], errors="coerce")
    df["descripcion"] = df.get("descripcion", pd.Series(dtype="string")).astype("string").fillna("")

    # marca cuota
    df["es_cuota"] = df["cuota"].notna() & df["fecha_cuota"].notna()
    df["estado"] = np.where(
        df["es_cuota"] & (df["fecha_cuota"] <= fecha_hoy) & (df.get("abono").fillna(0) > 0),
        "pagada",
        np.where(df["es_cuota"] & (df["fecha_cuota"] <= fecha_hoy), "vencida_sin_pago",
                 np.where(df["es_cuota"], "pendiente_futura", "concepto_inicial"))
    )
    _save(df, "fact_deuda_jeison.csv")
    return df


# =========================================================
# ORQUESTADOR
# =========================================================
def run(fecha_hoy: pd.Timestamp | None = None) -> dict:
    if fecha_hoy is None:
        fecha_hoy = pd.Timestamp.today().normalize()
    log.info("=== NORMALIZE (fecha_ref=%s) ===", fecha_hoy.date())
    results: dict[str, dict] = {}
    for name, fn in [
        ("ventas",         lambda: normalize_ventas()),
        ("inventario",     lambda: normalize_inventario()),
        ("cartera",        lambda: normalize_cartera(fecha_hoy)),
        ("cxp_nacional",   lambda: normalize_cxp_nacional(fecha_hoy)),
        ("cxp_internacional", lambda: normalize_cxp_internacional(fecha_hoy)),
        ("gastos",         lambda: normalize_gastos()),
        ("ventas_vendedor",lambda: normalize_ventas_vendedor()),
        ("pagos_jeison",   lambda: normalize_pagos_jeison(fecha_hoy)),
    ]:
        try:
            df = fn()
            results[name] = {"ok": True, "rows": len(df), "cols": len(df.columns)}
        except Exception as e:
            log.exception("Falla normalize %s: %s", name, e)
            results[name] = {"ok": False, "error": str(e)}
    return results


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    run()
