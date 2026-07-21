"""
Corrección de costos por juego de inventarios prorrateado.

Regla:
1. Extrae del ER (Gastos Global) el Costo de Ventas oficial por mes (Oct 2025 - Jun 2026).
2. Para meses no cubiertos por el ER (Jun 2024 - Sep 2025) usa el % promedio COGS/Ventas del ER.
3. Para cada línea:
   - Si costo_promedio > 0 en el dato original → RESPETAR (fuente_costo = "SISTEMA")
   - Si costo_promedio = 0/null → prorratea el REMANENTE del COGS del mes entre líneas sin
     costo, proporcionalmente al peso de cada línea en las ventas sin costo del mes.
     fuente_costo = "JUEGO_INV"
4. Recalcula costo_total_linea, margen_bruto_linea, margen_bruto_pct.
5. Guarda fact_ventas.csv (sobrescribe) + costo_ventas_er.csv.
"""
from __future__ import annotations
import logging
import re

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from config import CLEAN_DIR
from normalize import _raw_path

log = logging.getLogger(__name__)


# =========================================================
# 1. LECTURA DEL ER
# =========================================================
def _leer_er_2025(ws) -> dict[str, dict]:
    """ER 2025: cols B,C,D = Oct, Nov, Dic (row 12 tiene headers 'Octubre','Noviembre','Diciembre')."""
    meses_cols = {"2025-10": 2, "2025-11": 3, "2025-12": 4}
    out = {}
    for ym, c in meses_cols.items():
        ventas       = ws.cell(row=19, column=c).value or 0
        devoluciones = ws.cell(row=20, column=c).value or 0
        ventas_netas = ws.cell(row=21, column=c).value or 0
        costo_ventas = ws.cell(row=22, column=c).value or 0
        utilidad_bruta = ws.cell(row=23, column=c).value or 0
        inv_ini      = ws.cell(row=13, column=c).value or 0
        compras      = ws.cell(row=14, column=c).value or 0
        inv_fin      = ws.cell(row=15, column=c).value or 0
        out[ym] = {
            "ventas_bruta_er": float(ventas),
            "devoluciones_er": float(devoluciones),
            "ventas_netas_er": float(ventas_netas),
            "costo_ventas_er": float(costo_ventas),
            "utilidad_bruta_er": float(utilidad_bruta),
            "inv_inicial_er": float(inv_ini),
            "compras_er": float(compras),
            "inv_final_er": float(inv_fin),
        }
    return out


def _leer_er_2026(ws) -> dict[str, dict]:
    """ER 2026: cols B..G = Ene..Jun. Julio+ están vacíos."""
    meses_cols = {f"2026-{m:02d}": 1 + m for m in range(1, 7)}  # ene→2, jun→7
    out = {}
    for ym, c in meses_cols.items():
        ventas       = ws.cell(row=19, column=c).value or 0
        devoluciones = ws.cell(row=20, column=c).value or 0
        ventas_netas = ws.cell(row=21, column=c).value or 0
        costo_ventas = ws.cell(row=22, column=c).value or 0
        utilidad_bruta = ws.cell(row=23, column=c).value or 0
        inv_ini      = ws.cell(row=12, column=c).value or 0
        compras      = ws.cell(row=13, column=c).value or 0
        bonif        = ws.cell(row=14, column=c).value or 0
        inv_fin      = ws.cell(row=15, column=c).value or 0
        # Filtrar meses con ventas=0 (aún no llenados)
        if not ventas or ventas == 0:
            continue
        out[ym] = {
            "ventas_bruta_er": float(ventas),
            "devoluciones_er": float(devoluciones),
            "ventas_netas_er": float(ventas_netas),
            "costo_ventas_er": float(costo_ventas),
            "utilidad_bruta_er": float(utilidad_bruta),
            "inv_inicial_er": float(inv_ini),
            "compras_er": float(compras),
            "bonificaciones_er": float(bonif),
            "inv_final_er": float(inv_fin),
        }
    return out


def cargar_cogs_er() -> pd.DataFrame:
    """Devuelve DF con columnas: anio_mes, ventas_bruta_er, costo_ventas_er, margen_pct_er, y demás.
    Guarda además data/clean/costo_ventas_er.csv como snapshot oficial."""
    p = _raw_path("gastos")
    wb = load_workbook(p, data_only=True)
    data_25 = _leer_er_2025(wb["ER 2025"])
    data_26 = _leer_er_2026(wb["ER 2026"])
    combined = {**data_25, **data_26}
    rows = []
    for ym, d in sorted(combined.items()):
        row = {"anio_mes": ym, **d}
        ventas = d["ventas_bruta_er"]
        cogs = d["costo_ventas_er"]
        row["margen_bruto_pct_er"] = (ventas - cogs) / ventas if ventas > 0 else np.nan
        rows.append(row)
    df = pd.DataFrame(rows)
    # Ratio de devoluciones por mes (para prorratear en fact_ventas)
    df["ratio_devoluciones"] = df["devoluciones_er"] / df["ventas_bruta_er"].replace(0, np.nan)
    dest = CLEAN_DIR / "costo_ventas_er.csv"
    df.to_csv(dest, index=False, encoding="utf-8")
    log.info("  → costo_ventas_er.csv  (%d meses del ER)", len(df))
    return df


# =========================================================
# 2. CORRECCIÓN DE fact_ventas.csv
# =========================================================
def corregir_costos() -> dict:
    """Aplica juego de inventarios sobre fact_ventas.csv (sobrescribe).
    Devuelve estadísticas de la corrección."""
    log.info("[COST-FIX] Corrigiendo costos por juego de inventarios prorrateado…")
    er = cargar_cogs_er()

    # % promedio ponderado del ER (para meses sin ER)
    total_vbrutas_er = er["ventas_bruta_er"].sum()
    total_vnetas_er  = er["ventas_netas_er"].sum()
    total_cogs_er    = er["costo_ventas_er"].sum()
    total_dev_er     = er["devoluciones_er"].sum()
    # COGS/Ventas se calcula sobre Ventas NETAS (base oficial del margen)
    ratio_cogs_promedio = total_cogs_er / total_vnetas_er if total_vnetas_er > 0 else 0.83
    # Ratio de devoluciones promedio (para meses sin ER)
    ratio_dev_promedio = total_dev_er / total_vbrutas_er if total_vbrutas_er > 0 else 0.024
    log.info("  Ratio COGS/Ventas NETAS promedio del ER: %.4f", ratio_cogs_promedio)
    log.info("  Ratio Devoluciones/Ventas Brutas promedio: %.4f", ratio_dev_promedio)

    # Ventas de fact_ventas por mes
    fv = pd.read_csv(CLEAN_DIR / "fact_ventas.csv", low_memory=False)
    n_original = len(fv)
    fv["fecha"] = pd.to_datetime(fv["fecha"], errors="coerce")
    fv["anio_mes"] = fv["fecha"].dt.to_period("M").astype("string")

    # Marca si tenía costo original
    fv["tenia_costo_sistema"] = fv["costo_promedio"].fillna(0) > 0

    # Agregado por mes: ventas totales, ventas con costo, ventas sin costo, costos_sist
    agg = fv.groupby("anio_mes").agg(
        ventas_total_mes=("total_venta", "sum"),
        ventas_con_costo_mes=("total_venta", lambda s: s[fv.loc[s.index, "tenia_costo_sistema"]].sum()),
        ventas_sin_costo_mes=("total_venta", lambda s: s[~fv.loc[s.index, "tenia_costo_sistema"]].sum()),
        costo_sistema_mes=("cantidad", lambda s: (s * fv.loc[s.index, "costo_promedio"].fillna(0)).sum()),
    ).reset_index()

    # Merge con ER: si no está, estimar cogs y devoluciones con ratios promedio
    er_min = er[["anio_mes", "costo_ventas_er", "ventas_bruta_er",
                 "ventas_netas_er", "devoluciones_er", "ratio_devoluciones"]].copy()
    agg = agg.merge(er_min, on="anio_mes", how="left")
    agg["fuente_cogs"] = np.where(agg["costo_ventas_er"].notna(), "ER", "ESTIMADO")

    # Ventas Netas objetivo por mes:
    #   - Si hay ER: usar ventas_netas_er (fuerza que el neto agregado del fact cuadre con
    #     el reporte oficial, incluso si fact_ventas > ventas_bruta_er por ajustes).
    #   - Si no hay ER: aplicar ratio de devoluciones promedio.
    agg["ventas_netas_mes"] = np.where(
        agg["ventas_netas_er"].notna(),
        agg["ventas_netas_er"],
        agg["ventas_total_mes"] * (1 - ratio_dev_promedio),
    )
    # Devoluciones = brutas − netas
    agg["devoluciones_mes"] = (agg["ventas_total_mes"] - agg["ventas_netas_mes"]).clip(lower=0)
    # Ratio de devoluciones aplicable a cada línea del mes
    agg["ratio_dev_mes"] = np.where(
        agg["ventas_total_mes"] > 0,
        agg["devoluciones_mes"] / agg["ventas_total_mes"], 0.0,
    )

    # COGS del mes: del ER si está, si no estimarlo sobre Ventas Netas (base oficial)
    agg["costo_ventas_final"] = np.where(
        agg["costo_ventas_er"].notna(),
        agg["costo_ventas_er"],
        agg["ventas_netas_mes"] * ratio_cogs_promedio,
    )
    # Remanente COGS a prorratear entre líneas SIN costo del sistema
    agg["cogs_a_prorratear"] = (agg["costo_ventas_final"] - agg["costo_sistema_mes"]).round(2)
    # Prorrateo sobre VENTAS BRUTAS sin costo (mantenemos el criterio: la línea
    # aporta al COGS proporcional a su venta bruta)
    agg["ratio_prorrateo"] = np.where(
        agg["ventas_sin_costo_mes"] > 0,
        agg["cogs_a_prorratear"] / agg["ventas_sin_costo_mes"],
        0.0,
    )
    # Cap: si remanente < 0 forzar a 0 y reportar anomalía
    anomalias_neg = agg[agg["cogs_a_prorratear"] < 0].copy()
    if len(anomalias_neg) > 0:
        log.warning("  ⚠ %d meses tienen costo_sistema > COGS_ER — prorrateo forzado a 0:",
                    len(anomalias_neg))
        for _, r in anomalias_neg.iterrows():
            log.warning("    %s : sist=$%s cogs_er=$%s exceso=$%s",
                        r["anio_mes"], f"{r['costo_sistema_mes']:,.0f}",
                        f"{r['costo_ventas_final']:,.0f}", f"{-r['cogs_a_prorratear']:,.0f}")
    agg.loc[agg["cogs_a_prorratear"] < 0, "ratio_prorrateo"] = 0.0
    agg.loc[agg["cogs_a_prorratear"] < 0, "cogs_a_prorratear"] = 0.0

    # Aplicar a fact_ventas
    ratio_por_mes = dict(zip(agg["anio_mes"], agg["ratio_prorrateo"]))
    ratio_dev_por_mes = dict(zip(agg["anio_mes"], agg["ratio_dev_mes"]))
    fv["_ratio_cogs"] = fv["anio_mes"].map(ratio_por_mes).fillna(0.0)
    fv["_ratio_dev"]  = fv["anio_mes"].map(ratio_dev_por_mes).fillna(0.0)

    # Recalcular costo_total_linea (sobre venta bruta = criterio del prorrateo)
    fv["fuente_costo"] = np.where(fv["tenia_costo_sistema"], "SISTEMA", "JUEGO_INV")
    fv["costo_total_linea"] = np.where(
        fv["tenia_costo_sistema"],
        (fv["cantidad"] * fv["costo_promedio"].fillna(0)).round(2),
        (fv["total_venta"] * fv["_ratio_cogs"]).round(2),
    )
    # costo unitario final
    fv["costo_unitario_final"] = np.where(
        fv["cantidad"] > 0,
        (fv["costo_total_linea"] / fv["cantidad"]).round(2),
        np.nan,
    )

    # Devolución prorrateada y VENTA NETA por línea (base oficial del margen)
    fv["devolucion_prorrateada"] = (fv["total_venta"] * fv["_ratio_dev"]).round(2)
    fv["venta_neta_linea"]       = (fv["total_venta"] - fv["devolucion_prorrateada"]).round(2)

    # Margen recalculado sobre VENTAS NETAS (coherente con reporte oficial)
    fv["margen_bruto_linea"] = (fv["venta_neta_linea"] - fv["costo_total_linea"]).round(2)
    fv["margen_bruto_pct"] = np.where(
        fv["venta_neta_linea"] > 0,
        (fv["margen_bruto_linea"] / fv["venta_neta_linea"]).round(4),
        np.nan,
    )

    # Limpieza cols auxiliares
    fv = fv.drop(columns=["_ratio_cogs", "_ratio_dev", "tenia_costo_sistema"])

    # Guardar (sobrescribe)
    fv.to_csv(CLEAN_DIR / "fact_ventas.csv", index=False, encoding="utf-8")
    log.info("  → fact_ventas.csv actualizado (%d filas)", len(fv))

    # Estadísticas
    n_sistema  = int((fv["fuente_costo"] == "SISTEMA").sum())
    n_juego    = int((fv["fuente_costo"] == "JUEGO_INV").sum())
    ventas_brutas_tot = float(fv["total_venta"].sum())
    ventas_netas_tot  = float(fv["venta_neta_linea"].sum())
    costo_tot         = float(fv["costo_total_linea"].sum())
    ub_tot            = ventas_netas_tot - costo_tot
    margen_glob_pct   = ub_tot / ventas_netas_tot if ventas_netas_tot > 0 else 0

    stats = {
        "n_lineas": len(fv),
        "n_sistema": n_sistema,
        "n_juego_inv": n_juego,
        "pct_sistema": n_sistema / len(fv),
        "pct_juego_inv": n_juego / len(fv),
        "meses_er": int(agg["fuente_cogs"].eq("ER").sum()),
        "meses_estimado": int(agg["fuente_cogs"].eq("ESTIMADO").sum()),
        "meses_anomalia_negativa": int(len(anomalias_neg)),
        "ratio_cogs_promedio_er": float(ratio_cogs_promedio),
        "ratio_dev_promedio_er": float(ratio_dev_promedio),
        "ventas_brutas_total": ventas_brutas_tot,
        "ventas_netas_total":  ventas_netas_tot,
        "costo_total":         costo_tot,
        "utilidad_bruta_total": ub_tot,
        "margen_global_pct":    float(margen_glob_pct),
    }
    log.info("  Líneas SISTEMA: %d (%.1f%%) | JUEGO_INV: %d (%.1f%%)",
             n_sistema, stats["pct_sistema"]*100, n_juego, stats["pct_juego_inv"]*100)
    log.info("  Ventas Brutas: $%s | Devoluciones: $%s | Ventas Netas: $%s",
             f"{ventas_brutas_tot:,.0f}", f"{ventas_brutas_tot-ventas_netas_tot:,.0f}", f"{ventas_netas_tot:,.0f}")
    log.info("  Margen global CORREGIDO (sobre Ventas Netas): %.2f%%", margen_glob_pct * 100)
    return {"stats": stats, "agg": agg}


def run() -> dict:
    log.info("=== CORRECT COSTS ===")
    return corregir_costos()


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    run()
